import sys
import sqlite3
import os
import shutil
import uuid
from datetime import date, datetime
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QSize, QDate, QTimer
from PyQt5.QtGui import QFont, QIcon, QColor, QPalette, QPixmap, QIntValidator, QDoubleValidator


class SortableTableItem(QTableWidgetItem):
    """QTableWidgetItem 子类，支持数值/日期正确排序"""
    def __lt__(self, other):
        my = self.data(Qt.UserRole)
        ot = other.data(Qt.UserRole)
        if my is not None and ot is not None:
            return my < ot
        return super().__lt__(other)

# ===================== 全局样式 =====================
STYLE_SHEET = """
QMainWindow {
    background-color: #f5f0eb;
}
QWidget {
    font-family: "Microsoft YaHei", "PingFang SC", "Helvetica Neue", sans-serif;
}

/* ---- 登录页 ---- */
#loginCard {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #ffffff, stop:1 #faf7f4);
    border-radius: 18px;
    border: 1px solid #e8e0d8;
}
#loginTitle {
    font-size: 22px;
    font-weight: bold;
    color: #3a3a3a;
}
#loginSubtitle {
    font-size: 12px;
    color: #999;
}

/* ---- 通用输入框 ---- */
QLineEdit {
    border: 1px solid #d4ccc4;
    border-radius: 8px;
    padding: 6px 10px;
    font-size: 13px;
    background: #fff;
    color: #333;
}
QLineEdit:focus {
    border: 2px solid #b08968;
}
QLineEdit:read-only {
    background: #f5f0eb;
    color: #888;
}

/* ---- 按钮 ---- */
QPushButton {
    border: none;
    border-radius: 8px;
    padding: 8px 18px;
    font-size: 13px;
    font-weight: bold;
    color: #fff;
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #7a5c3e, stop:1 #96754e);
}
QPushButton:hover {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #6b4f33, stop:1 #866740);
}
QPushButton:pressed {
    background: #5a4028;
}
QPushButton#dangerBtn {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #d96b6b, stop:1 #e88a8a);
    color: #fff;
}
QPushButton#dangerBtn:hover {
    background: #c55050;
}
QPushButton#secondaryBtn {
    background: #e8e0d8;
    color: #6b5b4e;
}
QPushButton#secondaryBtn:hover {
    background: #d9cfc5;
}
QPushButton#returnBtn {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #3d8455, stop:1 #5a9e6f);
    color: #fff;
    min-height: 22px;
    font-size: 14px;
}
QPushButton#returnBtn:hover {
    background: #347348;
}
QPushButton#compensateBtn {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #b07820, stop:1 #c88a30);
    color: #fff;
    min-height: 22px;
    font-size: 14px;
}
QPushButton#compensateBtn:hover {
    background: #9a6818;
}
QPushButton#stockInBtn {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #3d8455, stop:1 #5a9e6f);
    min-height: 22px;
    font-size: 14px;
    color: #fff;
}
QPushButton#stockInBtn:hover {
    background: #347348;
}
QPushButton#stockOutBtn {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #c55050, stop:1 #d96b6b);
    min-height: 22px;
    font-size: 14px;
    color: #fff;
}
QPushButton#stockOutBtn:hover {
    background: #b04040;
}

/* ---- 左侧面板 ---- */
#stockLeftPanel {
    background: #fdfbf9;
    border-radius: 12px;
    border: 1px solid #e8e0d8;
}
#stockLeftPanel QLabel {
    border: none;
    background: transparent;
}
#stockLeftPanel QPushButton {
    border: none;
}
#clothingLeftPanel {
    background: #fdfbf9;
    border-radius: 12px;
    border: 1px solid #e8e0d8;
}
#clothingLeftPanel QLabel {
    border: none;
    background: transparent;
}
#clothingLeftPanel QPushButton {
    border: none;
}

/* ---- 日期选择器 ---- */
QDateEdit {
    border: 1px solid #d4ccc4;
    border-radius: 8px;
    padding: 6px 10px;
    font-size: 13px;
    background: #fff;
    color: #333;
}
QDateEdit:focus {
    border: 2px solid #b08968;
}
QDateEdit::drop-down {
    border: none;
    width: 24px;
}

/* ---- 表格 ---- */
QTableWidget {
    background: #fff;
    border: 1px solid #e0d8d0;
    border-radius: 10px;
    gridline-color: #f0ebe6;
    font-size: 13px;
    selection-background-color: #f5ece3;
    selection-color: #3a3a3a;
}
QTableWidget::item {
    padding: 5px 8px;
    border-bottom: 1px solid #f0ebe6;
}
QHeaderView::section {
    background: #f9f5f1;
    color: #6b5b4e;
    font-weight: bold;
    font-size: 12px;
    padding: 7px 10px;
    border: none;
    border-bottom: 2px solid #e0d8d0;
}
QHeaderView::section:hover {
    background: #f0e8e0;
    color: #b08968;
}

/* ---- 选项卡 ---- */
QTabWidget::pane {
    border: 1px solid #e0d8d0;
    border-radius: 10px;
    background: #fff;
    top: -1px;
}
QTabBar::tab {
    background: #e8e0d8;
    color: #6b5b4e;
    font-size: 14px;
    font-weight: bold;
    padding: 10px 40px;
    margin-right: 4px;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
    min-width: 140px;
}
QTabBar::tab:selected {
    background: #fff;
    color: #b08968;
    border-bottom: 3px solid #b08968;
}
QTabBar::tab:hover:!selected {
    background: #f0e8e0;
}

/* ---- 下拉框 ---- */
QComboBox {
    border: 1px solid #d4ccc4;
    border-radius: 8px;
    padding: 6px 10px;
    font-size: 13px;
    background: #fff;
    color: #333;
}
QComboBox:focus {
    border: 2px solid #b08968;
}
QComboBox::drop-down {
    border: none;
    width: 28px;
}
QComboBox QAbstractItemView {
    background: #fff;
    border: 1px solid #d4ccc4;
    border-radius: 6px;
    color: #333;
    selection-background-color: #f5ece3;
    selection-color: #333;
    outline: none;
    padding: 4px;
}
QComboBox QAbstractItemView::item {
    min-height: 28px;
    padding: 4px 8px;
    color: #333;
}
QComboBox QAbstractItemView::item:hover {
    background: #f5ece3;
    color: #3a3a3a;
}

/* ---- 分组框 ---- */
QGroupBox {
    font-size: 14px;
    font-weight: bold;
    color: #6b5b4e;
    border: 1px solid #e0d8d0;
    border-radius: 10px;
    margin-top: 14px;
    padding-top: 18px;
    background: #fdfbf9;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 16px;
    padding: 0 8px;
}

/* ---- 标签 ---- */
QLabel {
    color: #555;
    font-size: 13px;
}
QLabel#sectionTitle {
    font-size: 15px;
    font-weight: bold;
    color: #6b5b4e;
    border: none;
}
QLabel#statNumber {
    font-size: 28px;
    font-weight: bold;
    color: #b08968;
    border: none;
}
QLabel#statLabel {
    font-size: 11px;
    color: #999;
    border: none;
}

/* ---- 搜索框 ---- */
QLineEdit#searchInput {
    border: 1px solid #d4ccc4;
    border-radius: 20px;
    padding: 8px 16px;
    font-size: 13px;
    background: #fff;
}

/* ---- 消息框 & 对话框 ---- */
QMessageBox {
    background: #fdfbf9;
    border: 1px solid #e0d8d0;
    border-radius: 12px;
}
QMessageBox QLabel {
    font-size: 13px;
    color: #333;
    padding: 8px 4px;
    background: transparent;
}
QMessageBox QPushButton {
    min-width: 80px;
    min-height: 28px;
    padding: 6px 20px;
    font-size: 13px;
    font-weight: bold;
    border-radius: 8px;
    color: #fff;
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #7a5c3e, stop:1 #96754e);
}
QMessageBox QPushButton:hover {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #6b4f33, stop:1 #866740);
}
QMessageBox QPushButton:pressed {
    background: #5a4028;
}

QInputDialog {
    background: #fdfbf9;
    border: 1px solid #e0d8d0;
}
QInputDialog QLabel {
    font-size: 13px;
    color: #333;
    padding: 4px 2px;
    background: transparent;
}
QInputDialog QSpinBox, QInputDialog QDoubleSpinBox {
    border: 1px solid #d4ccc4;
    border-radius: 8px;
    padding: 6px 10px;
    font-size: 13px;
    background: #fff;
    color: #333;
}
QInputDialog QSpinBox:focus, QInputDialog QDoubleSpinBox:focus {
    border: 2px solid #b08968;
}
QInputDialog QPushButton {
    min-width: 80px;
    min-height: 28px;
    padding: 6px 20px;
    font-size: 13px;
    font-weight: bold;
    border-radius: 8px;
    color: #fff;
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #7a5c3e, stop:1 #96754e);
}
QInputDialog QPushButton:hover {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #6b4f33, stop:1 #866740);
}

QDialog {
    background: #fdfbf9;
    border: 1px solid #e0d8d0;
}
QDialog QLabel {
    color: #333;
    font-size: 13px;
    background: transparent;
}
QDialog QPushButton {
    min-width: 80px;
    min-height: 28px;
    padding: 6px 20px;
    font-size: 13px;
    font-weight: bold;
    border-radius: 8px;
    color: #fff;
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #7a5c3e, stop:1 #96754e);
}
QDialog QPushButton:hover {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #6b4f33, stop:1 #866740);
}

/* ---- 滚动条 ---- */
QScrollBar:vertical {
    background: #f5f0eb;
    width: 8px;
    border-radius: 4px;
    margin: 2px;
}
QScrollBar::handle:vertical {
    background: #d4ccc4;
    border-radius: 4px;
    min-height: 30px;
}
QScrollBar::handle:vertical:hover {
    background: #b08968;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}
QScrollBar:horizontal {
    background: #f5f0eb;
    height: 8px;
    border-radius: 4px;
    margin: 2px;
}
QScrollBar::handle:horizontal {
    background: #d4ccc4;
    border-radius: 4px;
    min-width: 30px;
}
QScrollBar::handle:horizontal:hover {
    background: #b08968;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0px;
}

/* ---- 工具提示 ---- */
QToolTip {
    background: #fdfbf9;
    color: #333;
    border: 1px solid #d4ccc4;
    border-radius: 6px;
    padding: 6px 10px;
    font-size: 12px;
}

/* ---- 日历弹窗 ---- */
QCalendarWidget {
    background: #fdfbf9;
    border: 1px solid #e0d8d0;
    border-radius: 8px;
}
QCalendarWidget QToolButton {
    color: #6b5b4e;
    font-size: 13px;
    font-weight: bold;
    background: transparent;
    border: none;
    padding: 4px 8px;
    border-radius: 6px;
}
QCalendarWidget QToolButton:hover {
    background: #f0e8e0;
}
QCalendarWidget QWidget#qt_calendar_navigationbar {
    background: #f9f5f1;
    border-bottom: 1px solid #e0d8d0;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
    padding: 4px;
}
QCalendarWidget QAbstractItemView {
    background: #fff;
    selection-background-color: #b08968;
    selection-color: #fff;
    font-size: 12px;
    outline: none;
}
QCalendarWidget QAbstractItemView:enabled {
    color: #333;
}
QCalendarWidget QAbstractItemView:disabled {
    color: #bbb;
}

/* ---- SpinBox 通用 ---- */
QSpinBox, QDoubleSpinBox {
    border: 1px solid #d4ccc4;
    border-radius: 8px;
    padding: 6px 10px;
    font-size: 13px;
    background: #fff;
    color: #333;
}
QSpinBox:focus, QDoubleSpinBox:focus {
    border: 2px solid #b08968;
}
QSpinBox::up-button, QDoubleSpinBox::up-button,
QSpinBox::down-button, QDoubleSpinBox::down-button {
    border: none;
    width: 18px;
}

/* ---- 进度条 ---- */
QProgressBar {
    border: 1px solid #d4ccc4;
    border-radius: 6px;
    text-align: center;
    font-size: 11px;
    color: #6b5b4e;
    background: #f5f0eb;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #b08968, stop:1 #c8a47e);
    border-radius: 5px;
}
"""

# ===================== 数据库操作类 =====================
class Database:
    def __init__(self):
        self.conn = sqlite3.connect('clothing_db.db')
        self.cursor = self.conn.cursor()
        self.create_tables()

    def create_tables(self):
        # 用户表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                password TEXT NOT NULL
            )
        ''')
        # 服装商品表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS clothing (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT DEFAULT '',
                name TEXT NOT NULL,
                category TEXT,
                brand TEXT,
                size TEXT,
                color TEXT,
                season TEXT,
                stock INTEGER DEFAULT 0,
                cost_price REAL DEFAULT 0
            )
        ''')
        # 兼容旧数据库：如果 clothing 表没有 code 列则添加
        try:
            self.cursor.execute("SELECT code FROM clothing LIMIT 1")
        except sqlite3.OperationalError:
            self.cursor.execute("ALTER TABLE clothing ADD COLUMN code TEXT DEFAULT ''")
        # 出入库记录表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS stock_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clothing_id INTEGER NOT NULL,
                clothing_name TEXT,
                department TEXT DEFAULT '',
                team_name TEXT DEFAULT '',
                host_operator TEXT DEFAULT '',
                anchor_name TEXT DEFAULT '',
                contact TEXT DEFAULT '',
                record_date TEXT,
                quantity INTEGER DEFAULT 0,
                direction INTEGER DEFAULT -1,
                status TEXT DEFAULT '借出',
                compensate_amount REAL DEFAULT 0,
                remark TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                return_date TEXT DEFAULT '',
                size TEXT DEFAULT '',
                image_path TEXT DEFAULT '',
                FOREIGN KEY (clothing_id) REFERENCES clothing(id)
            )
        ''')
        # 兼容旧数据库：如果 stock_records 表没有 return_date 列则添加
        try:
            self.cursor.execute("SELECT return_date FROM stock_records LIMIT 1")
        except sqlite3.OperationalError:
            self.cursor.execute("ALTER TABLE stock_records ADD COLUMN return_date TEXT DEFAULT ''")
        # 兼容旧数据库：如果 stock_records 表没有 size 列则添加
        try:
            self.cursor.execute("SELECT size FROM stock_records LIMIT 1")
        except sqlite3.OperationalError:
            self.cursor.execute("ALTER TABLE stock_records ADD COLUMN size TEXT DEFAULT ''")
        # 兼容旧数据库：如果 stock_records 表没有 image_path 列则添加
        try:
            self.cursor.execute("SELECT image_path FROM stock_records LIMIT 1")
        except sqlite3.OperationalError:
            self.cursor.execute("ALTER TABLE stock_records ADD COLUMN image_path TEXT DEFAULT ''")
        # 兼容旧数据库：如果 clothing 表没有 image_path 列则添加
        try:
            self.cursor.execute("SELECT image_path FROM clothing LIMIT 1")
        except sqlite3.OperationalError:
            self.cursor.execute("ALTER TABLE clothing ADD COLUMN image_path TEXT DEFAULT ''")
        # 兼容旧数据库：如果 users 表没有 expires_at 列则添加
        try:
            self.cursor.execute("SELECT expires_at FROM users LIMIT 1")
        except sqlite3.OperationalError:
            self.cursor.execute("ALTER TABLE users ADD COLUMN expires_at TEXT DEFAULT NULL")
        # 默认管理员（试用账号，30天有效）
        self.cursor.execute("SELECT * FROM users WHERE username='admin'")
        if not self.cursor.fetchone():
            exp = (datetime.now() + __import__('datetime').timedelta(days=30)).strftime('%Y-%m-%d')
            self.cursor.execute(
                "INSERT INTO users (username,password,expires_at) VALUES (?,?,?)",
                ("admin", "123456", exp))
        # 正式账号（长期有效）
        self.cursor.execute("SELECT * FROM users WHERE username='nicloth'")
        if not self.cursor.fetchone():
            self.cursor.execute(
                "INSERT INTO users (username,password,expires_at) VALUES (?,?,?)",
                ("nicloth", "nicloth2026", None))
        self.conn.commit()

    def check_login(self, username, password):
        self.cursor.execute(
            "SELECT username, expires_at FROM users WHERE username=? AND password=?",
            (username, password))
        return self.cursor.fetchone()

    def add_clothing(self, code, name, category, brand, size, color, season, stock, cost_price, image_path=''):
        self.cursor.execute(
            "INSERT INTO clothing (code,name,category,brand,size,color,season,stock,cost_price,image_path) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (code, name, category, brand, size, color, season, stock, cost_price, image_path))
        self.conn.commit()

    def update_clothing(self, id_, code, name, category, brand, size, color, season, stock, cost_price, image_path=''):
        self.cursor.execute(
            "UPDATE clothing SET code=?,name=?,category=?,brand=?,size=?,color=?,season=?,stock=?,cost_price=?,image_path=? WHERE id=?",
            (code, name, category, brand, size, color, season, stock, cost_price, image_path, id_))
        self.conn.commit()

    def delete_clothing(self, id_):
        self.cursor.execute("DELETE FROM clothing WHERE id=?", (id_,))
        self.conn.commit()

    def get_all_clothing(self):
        self.cursor.execute("SELECT id,code,name,category,brand,size,color,season,stock,cost_price,image_path FROM clothing")
        return self.cursor.fetchall()

    def search_clothing(self, keyword):
        query = "SELECT id,code,name,category,brand,size,color,season,stock,cost_price,image_path FROM clothing WHERE name LIKE ? OR code LIKE ? OR category LIKE ? OR brand LIKE ? OR color LIKE ?"
        pattern = f"%{keyword}%"
        self.cursor.execute(query, (pattern, pattern, pattern, pattern, pattern))
        return self.cursor.fetchall()

    def find_clothing_by_code(self, code, size=None):
        if size:
            self.cursor.execute("SELECT id, name, stock FROM clothing WHERE code=? AND size=?", (code, size))
        else:
            self.cursor.execute("SELECT id, name, stock FROM clothing WHERE code=?", (code,))
        return self.cursor.fetchone()

    def find_sizes_by_code(self, code):
        self.cursor.execute(
            "SELECT size, stock FROM clothing WHERE code=? AND stock > 0 ORDER BY size", (code,))
        return self.cursor.fetchall()

    def get_cost_price(self, clothing_id):
        self.cursor.execute("SELECT cost_price FROM clothing WHERE id=?", (clothing_id,))
        row = self.cursor.fetchone()
        return row[0] if row else 0

    def update_stock(self, id_, num):
        self.cursor.execute("SELECT stock FROM clothing WHERE id=?", (id_,))
        row = self.cursor.fetchone()
        if row is None:
            return False
        new_stock = row[0] + num
        if new_stock < 0:
            return False
        self.cursor.execute("UPDATE clothing SET stock=? WHERE id=?", (new_stock, id_))
        self.conn.commit()
        return True

    # ---- 出入库记录 ----
    def add_stock_record(self, clothing_id, clothing_name, department, team_name,
                         host_operator, anchor_name, contact, record_date,
                         quantity, direction, status, remark, size='', image_path=''):
        self.cursor.execute('''
            INSERT INTO stock_records
            (clothing_id, clothing_name, department, team_name, host_operator,
             anchor_name, contact, record_date, quantity, direction, status, remark, size, image_path)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (clothing_id, clothing_name, department, team_name, host_operator,
              anchor_name, contact, record_date, quantity, direction, status, remark, size, image_path))
        self.conn.commit()
        return self.cursor.lastrowid

    def get_all_stock_records(self):
        self.cursor.execute('''
            SELECT sr.id, COALESCE(c.code,''), sr.clothing_name, COALESCE(sr.size,''),
                   sr.department, sr.team_name,
                   sr.host_operator, sr.anchor_name, sr.contact, sr.record_date,
                   sr.quantity, sr.direction, sr.status,
                   sr.created_at, COALESCE(sr.return_date,''),
                   sr.remark, sr.compensate_amount, COALESCE(sr.image_path,'')
            FROM stock_records sr LEFT JOIN clothing c ON sr.clothing_id = c.id
            ORDER BY sr.id DESC
        ''')
        return self.cursor.fetchall()

    def search_stock_records(self, keyword):
        pattern = f"%{keyword}%"
        self.cursor.execute('''
            SELECT sr.id, COALESCE(c.code,''), sr.clothing_name, COALESCE(sr.size,''),
                   sr.department, sr.team_name,
                   sr.host_operator, sr.anchor_name, sr.contact, sr.record_date,
                   sr.quantity, sr.direction, sr.status,
                   sr.created_at, COALESCE(sr.return_date,''),
                   sr.remark, sr.compensate_amount, COALESCE(sr.image_path,'')
            FROM stock_records sr LEFT JOIN clothing c ON sr.clothing_id = c.id
            WHERE sr.clothing_name LIKE ? OR c.code LIKE ? OR sr.department LIKE ? OR sr.team_name LIKE ?
               OR sr.host_operator LIKE ? OR sr.anchor_name LIKE ? OR sr.contact LIKE ? OR sr.remark LIKE ?
               OR sr.size LIKE ?
            ORDER BY sr.id DESC
        ''', (pattern, pattern, pattern, pattern, pattern, pattern, pattern, pattern, pattern))
        return self.cursor.fetchall()

    def update_record_status(self, record_id, new_status, compensate_amount=0):
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        self.cursor.execute(
            "UPDATE stock_records SET status=?, compensate_amount=?, return_date=? WHERE id=?",
            (new_status, compensate_amount, now, record_id))
        self.conn.commit()

    def update_record_quantity(self, record_id, new_quantity):
        self.cursor.execute(
            "UPDATE stock_records SET quantity=? WHERE id=?",
            (new_quantity, record_id))
        self.conn.commit()

    def split_record_for_return(self, record_id, return_qty):
        """将借出记录拆分：return_qty 件标记为已归还，剩余保持借出"""
        record = self.get_record_by_id(record_id)
        if not record:
            return False
        total_qty = record[9]
        if return_qty > total_qty or return_qty <= 0:
            return False
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        if return_qty == total_qty:
            # 全部归还
            self.cursor.execute(
                "UPDATE stock_records SET status='已归还', return_date=? WHERE id=?",
                (now, record_id))
        else:
            # 减少原记录数量
            self.cursor.execute(
                "UPDATE stock_records SET quantity=? WHERE id=?",
                (total_qty - return_qty, record_id))
            # 新建一条已归还记录
            self.cursor.execute('''
                INSERT INTO stock_records
                (clothing_id, clothing_name, department, team_name, host_operator,
                 anchor_name, contact, record_date, quantity, direction, status, remark,
                 created_at, return_date, size, image_path)
                SELECT clothing_id, clothing_name, department, team_name, host_operator,
                       anchor_name, contact, record_date, ?, direction, '已归还', remark,
                       created_at, ?, size, image_path
                FROM stock_records WHERE id=?
            ''', (return_qty, now, record_id))
        self.conn.commit()
        return True

    def split_record_for_compensate(self, record_id, comp_qty, comp_amount):
        """将借出记录拆分：comp_qty 件标记为已赔付，剩余保持借出"""
        record = self.get_record_by_id(record_id)
        if not record:
            return False
        total_qty = record[9]
        if comp_qty > total_qty or comp_qty <= 0:
            return False
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        if comp_qty == total_qty:
            # 全部赔付
            self.cursor.execute(
                "UPDATE stock_records SET status='已赔付', compensate_amount=?, return_date=? WHERE id=?",
                (comp_amount, now, record_id))
        else:
            # 减少原记录数量
            self.cursor.execute(
                "UPDATE stock_records SET quantity=? WHERE id=?",
                (total_qty - comp_qty, record_id))
            # 新建一条已赔付记录
            self.cursor.execute('''
                INSERT INTO stock_records
                (clothing_id, clothing_name, department, team_name, host_operator,
                 anchor_name, contact, record_date, quantity, direction, status,
                 compensate_amount, remark, created_at, return_date, size, image_path)
                SELECT clothing_id, clothing_name, department, team_name, host_operator,
                       anchor_name, contact, record_date, ?, direction, '已赔付',
                       ?, remark, created_at, ?, size, image_path
                FROM stock_records WHERE id=?
            ''', (comp_qty, comp_amount, now, record_id))
        self.conn.commit()
        return True

    def get_record_by_id(self, record_id):
        self.cursor.execute('''
            SELECT id, clothing_id, clothing_name, department, team_name,
                   host_operator, anchor_name, contact, record_date,
                   quantity, direction, status, compensate_amount, remark,
                   COALESCE(size,''), COALESCE(image_path,'')
            FROM stock_records WHERE id=?
        ''', (record_id,))
        return self.cursor.fetchone()

    def update_record_fields(self, record_id, department, team_name, host_operator,
                             anchor_name, contact, record_date, quantity, remark,
                             size='', image_path=None):
        """更新出入库记录的可编辑字段，image_path=None 表示不修改图片"""
        if image_path is not None:
            self.cursor.execute('''
                UPDATE stock_records SET department=?, team_name=?, host_operator=?,
                       anchor_name=?, contact=?, record_date=?, quantity=?,
                       remark=?, size=?, image_path=?
                WHERE id=?
            ''', (department, team_name, host_operator, anchor_name, contact,
                  record_date, quantity, remark, size, image_path, record_id))
        else:
            self.cursor.execute('''
                UPDATE stock_records SET department=?, team_name=?, host_operator=?,
                       anchor_name=?, contact=?, record_date=?, quantity=?,
                       remark=?, size=?
                WHERE id=?
            ''', (department, team_name, host_operator, anchor_name, contact,
                  record_date, quantity, remark, size, record_id))
        self.conn.commit()

    def delete_stock_record(self, record_id):
        self.cursor.execute('DELETE FROM stock_records WHERE id=?', (record_id,))
        self.conn.commit()

    def get_stats(self):
        self.cursor.execute("SELECT COUNT(*), COALESCE(SUM(stock),0) FROM clothing")
        count, total_stock = self.cursor.fetchone()
        self.cursor.execute("SELECT COUNT(*) FROM clothing WHERE stock <= 5")
        low_stock = self.cursor.fetchone()[0]
        self.cursor.execute("SELECT COALESCE(SUM(quantity),0) FROM stock_records WHERE status='借出'")
        borrowed_qty = self.cursor.fetchone()[0]
        return count, total_stock, low_stock, borrowed_qty


# ===================== 登录窗口 =====================
class LoginWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Nicloth（服装管理系统）')
        self.setFixedSize(420, 480)
        self.setStyleSheet("background-color: #f5f0eb;")

        central = QWidget()
        self.setCentralWidget(central)
        outer = QVBoxLayout(central)
        outer.setAlignment(Qt.AlignCenter)

        # 登录卡片
        card = QWidget()
        card.setObjectName('loginCard')
        card.setFixedSize(340, 380)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(14)
        card_layout.setContentsMargins(36, 36, 36, 36)

        # Logo 区域
        logo_label = QLabel('👗')
        logo_label.setAlignment(Qt.AlignCenter)
        logo_label.setStyleSheet('font-size: 48px; border: none;')
        card_layout.addWidget(logo_label)

        title = QLabel('Nicloth（服装管理系统）')
        title.setObjectName('loginTitle')
        title.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(title)

        subtitle = QLabel('Clothing Inventory Management')
        subtitle.setObjectName('loginSubtitle')
        subtitle.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(subtitle)

        card_layout.addSpacing(10)

        # 账号
        user_label = QLabel('👤  账号')
        user_label.setStyleSheet('font-size: 12px; color: #888;')
        self.user_input = QLineEdit()
        self.user_input.setPlaceholderText('请输入账号')
        self.user_input.setText('admin')
        card_layout.addWidget(user_label)
        card_layout.addWidget(self.user_input)

        # 密码
        pwd_label = QLabel('🔒  密码')
        pwd_label.setStyleSheet('font-size: 12px; color: #888;')
        self.pwd_input = QLineEdit()
        self.pwd_input.setPlaceholderText('请输入密码')
        self.pwd_input.setEchoMode(QLineEdit.Password)
        self.pwd_input.setText('123456')
        card_layout.addWidget(pwd_label)
        card_layout.addWidget(self.pwd_input)

        card_layout.addSpacing(8)

        # 登录按钮
        self.login_btn = QPushButton('登  录')
        self.login_btn.setStyleSheet('''
            QPushButton {
                min-height: 36px; font-size: 16px; font-weight: bold;
                border-radius: 10px; color: #fff;
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #7a5c3e, stop:1 #96754e);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #6b4f33, stop:1 #866740);
            }
            QPushButton:pressed {
                background: #5a4028;
            }
        ''')
        self.login_btn.clicked.connect(self.login)
        self.pwd_input.returnPressed.connect(self.login)
        card_layout.addWidget(self.login_btn)

        card_layout.addStretch()
        outer.addWidget(card)

    def login(self):
        username = self.user_input.text().strip()
        password = self.pwd_input.text().strip()
        if not username or not password:
            QMessageBox.warning(self, '提示', '请输入账号和密码')
            return
        row = self.db.check_login(username, password)
        if row:
            expires_at = row[1]
            if expires_at:
                from datetime import datetime as _dt
                try:
                    exp_date = _dt.strptime(expires_at, '%Y-%m-%d').date()
                    today = _dt.now().date()
                    if today > exp_date:
                        QMessageBox.warning(
                            self, '账号已过期',
                            f'账号「{username}」已于 {expires_at} 过期\n'
                            '请联系管理员或使用其他账号登录')
                        return
                    remaining = (exp_date - today).days
                    if remaining <= 7:
                        QMessageBox.information(
                            self, '账号即将过期',
                            f'账号「{username}」将于 {expires_at} 过期\n'
                            f'剩余 {remaining} 天，请及时联系管理员续期')
                except ValueError:
                    pass
            self.main_win = MainWindow(username)
            self.main_win.show()
            self.close()
        else:
            QMessageBox.warning(self, '登录失败', '账号或密码错误，请重试')


# ===================== 主功能窗口 =====================
class MainWindow(QMainWindow):
    CATEGORIES = ['上衣', '裤装', '裙装', '外套', '内衣', '鞋履', '配饰', '其他']
    SIZES = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL', '均码']
    SEASONS = ['春季', '夏季', '秋季', '冬季', '四季通用']
    COLUMNS = ['', 'ID', '编号', '名称', '分类', '品牌', '尺码', '颜色', '季节', '库存', '进价(¥)', '图片']

    def __init__(self, username='admin'):
        super().__init__()
        self.db = Database()
        self.username = username
        self.init_ui()
        self.load_clothing()
        self.load_stock_records()
        self._setup_auto_backup()

    def init_ui(self):
        self.setWindowTitle('Nicloth（服装管理系统）')
        self.setMinimumSize(1080, 680)
        self.resize(1160, 720)

        # 主布局
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(20, 12, 20, 12)
        main_layout.setSpacing(10)

        # ---- 顶部栏 ----
        header = QHBoxLayout()
        app_title = QLabel('👗 Nicloth（服装管理系统）')
        app_title.setStyleSheet('font-size: 20px; font-weight: bold; color: #3a3a3a;')
        header.addWidget(app_title)
        header.addStretch()
        help_btn = QPushButton('📖 使用手册')
        help_btn.setObjectName('secondaryBtn')
        help_btn.setStyleSheet('QPushButton { padding: 5px 14px; font-size: 12px; }')
        help_btn.clicked.connect(self._show_help_manual)
        header.addWidget(help_btn)
        user_label = QLabel(f'👤 {self.username}')
        user_label.setStyleSheet('font-size: 13px; color: #888; margin-right: 8px;')
        header.addWidget(user_label)
        main_layout.addLayout(header)

        # ---- 统计卡片 ----
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(14)
        self.stat_total = self._make_stat_card('商品总数', '0', '#b08968')
        self.stat_stock = self._make_stat_card('总库存量', '0', '#5a9e6f')
        self.stat_low = self._make_stat_card('低库存预警', '0', '#d96b6b')
        self.stat_cats = self._make_stat_card('待归还', '0', '#c88a30')
        stats_layout.addWidget(self.stat_total)
        stats_layout.addWidget(self.stat_stock)
        stats_layout.addWidget(self.stat_low)
        stats_layout.addWidget(self.stat_cats)
        main_layout.addLayout(stats_layout)

        # ---- 选项卡 ----
        self.tab = QTabWidget()
        self.tab.addTab(self._create_stock_tab(), '🔄  出入库管理')
        self.tab.addTab(self._create_clothing_tab(), '📦  商品管理')
        self.tab.addTab(self._create_settings_tab(), '🛠  系统设置')
        main_layout.addWidget(self.tab, 1)

    def _update_sort_indicator(self, table, base_cols, sorted_col):
        """更新表头显示排序箭头 ▲/▼"""
        order = table.horizontalHeader().sortIndicatorOrder()
        arrow = ' ▲' if order == Qt.AscendingOrder else ' ▼'
        for i, name in enumerate(base_cols):
            if i == sorted_col:
                table.horizontalHeaderItem(i).setText(name + arrow)
            else:
                table.horizontalHeaderItem(i).setText(name)

    # ---------- 统计卡片 ----------
    def _make_stat_card(self, label_text, number, color):
        card = QWidget()
        card.setStyleSheet(f'''
            QWidget {{
                background: #fff;
                border-radius: 12px;
                border: 1px solid #eee;
            }}
        ''')
        card.setFixedHeight(88)
        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignCenter)
        layout.setSpacing(2)
        num = QLabel(number)
        num.setObjectName('statNumber')
        num.setAlignment(Qt.AlignCenter)
        num.setStyleSheet(f'font-size: 28px; font-weight: bold; color: {color}; border: none;')
        lbl = QLabel(label_text)
        lbl.setObjectName('statLabel')
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet('font-size: 11px; color: #999; border: none;')
        layout.addWidget(num)
        layout.addWidget(lbl)
        card._num_label = num
        return card

    def _update_stats(self):
        count, total_stock, low, borrowed = self.db.get_stats()
        self.stat_total._num_label.setText(str(count))
        self.stat_stock._num_label.setText(str(total_stock))
        self.stat_low._num_label.setText(str(low))
        self.stat_cats._num_label.setText(str(borrowed))

    # ---------- 内联校验工具 ----------
    def _wrap_field(self, widget):
        """将输入控件包装：控件 + 隐藏的错误提示 QLabel，返回 (container, error_label)"""
        container = QWidget()
        vbox = QVBoxLayout(container)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(2)
        vbox.addWidget(widget)
        err = QLabel()
        err.setStyleSheet('color: #e53935; font-size: 11px; padding: 0; margin: 0;')
        err.setWordWrap(True)
        err.hide()
        vbox.addWidget(err)
        if not hasattr(self, '_field_error_labels'):
            self._field_error_labels = {}
        self._field_error_labels[widget] = err
        return container

    def _set_field_error(self, widget, msg):
        """为控件设置红色边框 + 显示错误文字"""
        border_css = 'border: 1.5px solid #e53935; border-radius: 4px;'
        if isinstance(widget, QComboBox):
            widget.setStyleSheet(f'QComboBox {{ {border_css} }}')
        else:
            widget.setStyleSheet(border_css)
        lbl = self._field_error_labels.get(widget)
        if lbl:
            lbl.setText(msg)
            lbl.show()
        widget.setFocus()

    def _clear_field_error(self, widget):
        """清除单个控件的错误状态"""
        widget.setStyleSheet('')
        lbl = self._field_error_labels.get(widget)
        if lbl:
            lbl.hide()

    def _clear_all_field_errors(self):
        """清除所有已注册控件的错误状态"""
        for widget, lbl in getattr(self, '_field_error_labels', {}).items():
            widget.setStyleSheet('')
            lbl.hide()

    # ---------- 商品管理标签 ----------
    def _create_clothing_tab(self):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setSpacing(16)

        # ── 左侧表单 ──
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedWidth(320)
        scroll.setStyleSheet('QScrollArea { border: none; background: transparent; }')
        left_panel = QWidget()
        left_panel.setObjectName('clothingLeftPanel')
        form_outer = QVBoxLayout(left_panel)
        form_outer.setContentsMargins(18, 14, 18, 14)
        form_outer.setSpacing(8)

        form_title = QLabel('商品信息')
        form_title.setObjectName('sectionTitle')
        form_outer.addWidget(form_title)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        self.c_id = QLineEdit()
        self.c_id.setReadOnly(True)
        self.c_id.setPlaceholderText('自动生成')
        self.c_id.setStyleSheet('background: #f0ebe6; color: #999;')
        self.c_code = QLineEdit()
        self.c_code.setPlaceholderText('输入编号')
        self.c_name = QLineEdit()
        self.c_name.setPlaceholderText('输入服装名称')
        self.c_category = QComboBox()
        self.c_category.setEditable(True)
        self.c_category.addItems(self.CATEGORIES)
        self.c_category.setCurrentIndex(0)
        self.c_brand = QLineEdit()
        self.c_brand.setPlaceholderText('输入品牌')
        self.c_size = QComboBox()
        self.c_size.setEditable(True)
        self.c_size.addItems(self.SIZES)
        self.c_size.setCurrentIndex(0)
        self.c_color = QLineEdit()
        self.c_color.setPlaceholderText('输入颜色')
        self.c_season = QComboBox()
        self.c_season.setEditable(True)
        self.c_season.addItems(self.SEASONS)
        self.c_season.setCurrentIndex(0)
        self.c_stock = QLineEdit()
        self.c_stock.setPlaceholderText('0')
        self.c_stock.setValidator(QIntValidator(0, 999999, self))
        self.c_cost = QLineEdit()
        self.c_cost.setPlaceholderText('0.00')
        self.c_cost.setValidator(QDoubleValidator(0, 999999, 2, self))

        form.addRow('ID：', self.c_id)
        form.addRow('编号：', self.c_code)
        form.addRow('名称：', self._wrap_field(self.c_name))
        form.addRow('分类：', self.c_category)
        form.addRow('品牌：', self.c_brand)
        form.addRow('尺码：', self.c_size)
        form.addRow('颜色：', self.c_color)
        form.addRow('季节：', self.c_season)
        form.addRow('库存：', self._wrap_field(self.c_stock))
        form.addRow('进价(¥)：', self._wrap_field(self.c_cost))
        form_outer.addLayout(form)

        form_outer.addSpacing(8)

        # 图片上传
        self._clothing_image_path = ''
        c_img_btn_row = QHBoxLayout()
        c_img_btn_row.setSpacing(6)
        self.c_img_btn = QPushButton('📷 选择图片')
        self.c_img_btn.setObjectName('secondaryBtn')
        self.c_img_btn.clicked.connect(self._select_clothing_image)
        self.c_img_clear_btn = QPushButton('✕')
        self.c_img_clear_btn.setFixedWidth(30)
        self.c_img_clear_btn.setObjectName('secondaryBtn')
        self.c_img_clear_btn.clicked.connect(self._clear_clothing_image)
        c_img_btn_row.addWidget(self.c_img_btn)
        c_img_btn_row.addWidget(self.c_img_clear_btn)
        form_outer.addLayout(c_img_btn_row)

        self.c_img_preview = QLabel()
        self.c_img_preview.setFixedSize(280, 180)
        self.c_img_preview.setAlignment(Qt.AlignCenter)
        self.c_img_preview.setStyleSheet(
            'QLabel { border: 1px dashed #c4a882; border-radius: 6px; '
            'background: #faf7f4; color: #aaa; font-size: 12px; }')
        self.c_img_preview.setText('暂无图片')
        self.c_img_preview.setCursor(Qt.PointingHandCursor)
        self.c_img_preview.mousePressEvent = self._preview_clothing_image
        form_outer.addWidget(self.c_img_preview)

        form_outer.addSpacing(8)

        # 按钮
        btn_row1 = QHBoxLayout()
        self.add_btn = QPushButton('✚ 新增')
        self.add_btn.setObjectName('secondaryBtn')
        self.update_btn = QPushButton('✎ 修改')
        self.update_btn.setObjectName('secondaryBtn')
        btn_row1.addWidget(self.add_btn)
        btn_row1.addWidget(self.update_btn)
        form_outer.addLayout(btn_row1)

        btn_row2 = QHBoxLayout()
        self.del_btn = QPushButton('✕ 删除')
        self.del_btn.setObjectName('dangerBtn')
        self.clear_btn = QPushButton('↺ 清空')
        self.clear_btn.setObjectName('secondaryBtn')
        btn_row2.addWidget(self.del_btn)
        btn_row2.addWidget(self.clear_btn)
        form_outer.addLayout(btn_row2)

        self.add_btn.clicked.connect(self.add_clothing)
        self.update_btn.clicked.connect(self.update_clothing)
        self.del_btn.clicked.connect(self.delete_clothing)
        self.clear_btn.clicked.connect(self.clear_form)

        # Excel 导入/导出按钮
        btn_row3 = QHBoxLayout()
        self.import_btn = QPushButton('📥 Excel导入')
        self.import_btn.setObjectName('secondaryBtn')
        self.export_btn = QPushButton('📤 导出模板')
        self.export_btn.setObjectName('secondaryBtn')
        btn_row3.addWidget(self.import_btn)
        btn_row3.addWidget(self.export_btn)
        form_outer.addLayout(btn_row3)

        self.import_btn.clicked.connect(self.import_excel)
        self.export_btn.clicked.connect(self.export_template)

        form_outer.addStretch()

        scroll.setWidget(left_panel)

        # ── 右侧表格 ──
        right_panel = QVBoxLayout()
        right_panel.setSpacing(10)

        # 搜索栏
        search_row = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setObjectName('searchInput')
        self.search_input.setPlaceholderText('🔍  搜索名称 / 分类 / 品牌 / 颜色 ...')
        self.search_input.textChanged.connect(self.search_clothing)
        search_row.addWidget(self.search_input)
        right_panel.addLayout(search_row)

        # 表格
        self.clothing_table = QTableWidget()
        self.clothing_table.setColumnCount(len(self.COLUMNS))
        self.clothing_table.setHorizontalHeaderLabels(self.COLUMNS)
        self.clothing_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.clothing_table.setSelectionMode(QTableWidget.SingleSelection)
        self.clothing_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.clothing_table.setAlternatingRowColors(True)
        self.clothing_table.setStyleSheet('''
            QTableWidget { alternate-background-color: #faf7f4; }
        ''')
        self.clothing_table.horizontalHeader().setStretchLastSection(True)
        self.clothing_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.clothing_table.horizontalHeader().setMinimumSectionSize(60)
        self.clothing_table.setColumnWidth(0, 36)
        self.clothing_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.clothing_table.resizeColumnsToContents()
        self.clothing_table.verticalHeader().setVisible(False)
        self.clothing_table.setSortingEnabled(True)
        self.clothing_table.horizontalHeader().setSortIndicatorShown(False)
        self.clothing_table.horizontalHeader().sectionClicked.connect(
            lambda col: self._on_clothing_header_clicked(col))
        self.clothing_table.cellClicked.connect(self.select_clothing)
        right_panel.addWidget(self.clothing_table, 1)

        # ── 批量操作栏 ──
        batch_bar = QHBoxLayout()
        batch_bar.setSpacing(10)
        self.batch_select_all = QPushButton('☐ 全选')
        self.batch_select_all.setObjectName('secondaryBtn')
        self.batch_select_all.setMinimumWidth(70)
        self.batch_select_all.clicked.connect(self._toggle_select_all)
        batch_bar.addWidget(self.batch_select_all)
        self.batch_selected_label = QLabel('已选 0 项')
        self.batch_selected_label.setStyleSheet('color: #888; font-size: 12px;')
        batch_bar.addWidget(self.batch_selected_label)
        batch_bar.addStretch()
        batch_add_stock = QPushButton('▲ 批量加库存')
        batch_add_stock.setObjectName('secondaryBtn')
        batch_add_stock.clicked.connect(lambda: self._batch_adjust_stock(1))
        batch_sub_stock = QPushButton('▼ 批量减库存')
        batch_sub_stock.setObjectName('secondaryBtn')
        batch_sub_stock.clicked.connect(lambda: self._batch_adjust_stock(-1))
        batch_del = QPushButton('✕ 批量删除')
        batch_del.setObjectName('dangerBtn')
        batch_del.clicked.connect(self._batch_delete)
        batch_export = QPushButton('📊 导出Excel')
        batch_export.setObjectName('secondaryBtn')
        batch_export.clicked.connect(self._export_clothing_excel)
        batch_bar.addWidget(batch_export)
        batch_bar.addWidget(batch_add_stock)
        batch_bar.addWidget(batch_sub_stock)
        batch_bar.addWidget(batch_del)
        right_panel.addLayout(batch_bar)

        # ── 分页栏 ──
        self._clothing_page = 1
        self._clothing_page_size = 50
        self._clothing_data = []
        pager = QHBoxLayout()
        pager.setSpacing(8)
        self.cl_page_label = QLabel('共 0 条')
        self.cl_page_label.setStyleSheet('color: #888; font-size: 12px;')
        pager.addWidget(self.cl_page_label)
        pager.addStretch()
        self.cl_first_btn = QPushButton('«')
        self.cl_prev_btn = QPushButton('‹')
        self.cl_page_info = QLabel('1 / 1')
        self.cl_page_info.setStyleSheet('color: #6b5b4e; font-size: 12px; font-weight: bold;')
        self.cl_page_info.setAlignment(Qt.AlignCenter)
        self.cl_page_info.setMinimumWidth(60)
        self.cl_next_btn = QPushButton('›')
        self.cl_last_btn = QPushButton('»')
        for b in [self.cl_first_btn, self.cl_prev_btn, self.cl_next_btn, self.cl_last_btn]:
            b.setObjectName('secondaryBtn')
            b.setFixedSize(32, 28)
            b.setStyleSheet('QPushButton { padding: 0; font-size: 14px; }')
        self.cl_first_btn.clicked.connect(lambda: self._clothing_go_page(1))
        self.cl_prev_btn.clicked.connect(lambda: self._clothing_go_page(self._clothing_page - 1))
        self.cl_next_btn.clicked.connect(lambda: self._clothing_go_page(self._clothing_page + 1))
        self.cl_last_btn.clicked.connect(lambda: self._clothing_go_page(self._clothing_total_pages()))
        pager.addWidget(self.cl_first_btn)
        pager.addWidget(self.cl_prev_btn)
        pager.addWidget(self.cl_page_info)
        pager.addWidget(self.cl_next_btn)
        pager.addWidget(self.cl_last_btn)
        pager.addStretch()
        lbl = QLabel('每页')
        lbl.setStyleSheet('color: #888; font-size: 12px;')
        self.cl_page_size_combo = QComboBox()
        self.cl_page_size_combo.addItems(['20', '50', '100', '200'])
        self.cl_page_size_combo.setCurrentText('50')
        self.cl_page_size_combo.setFixedWidth(65)
        self.cl_page_size_combo.currentTextChanged.connect(self._clothing_page_size_changed)
        lbl2 = QLabel('条')
        lbl2.setStyleSheet('color: #888; font-size: 12px;')
        pager.addWidget(lbl)
        pager.addWidget(self.cl_page_size_combo)
        pager.addWidget(lbl2)
        right_panel.addLayout(pager)

        layout.addWidget(scroll)
        layout.addLayout(right_panel, 1)
        return widget

    # ---------- 出入库标签 ----------
    STOCK_RECORD_COLS = ['记录ID', '商品编号', '商品名称', '尺码', '部门', '团名',
                         '主持/运营', '主播艺名', '联系方式', '操作日期',
                         '数量', '类型', '状态', '借出时间', '归还时间',
                         '备注', '赔付金额(¥)', '图片']

    def _create_stock_tab(self):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setSpacing(16)

        # ── 左侧表单 ──
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedWidth(340)
        scroll.setStyleSheet('QScrollArea { border: none; background: transparent; }')
        left_panel = QWidget()
        left_panel.setObjectName('stockLeftPanel')
        form_outer = QVBoxLayout(left_panel)
        form_outer.setContentsMargins(18, 14, 18, 14)
        form_outer.setSpacing(8)

        form_title = QLabel('出入库操作')
        form_title.setObjectName('sectionTitle')
        form_outer.addWidget(form_title)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        self.stock_id = QLineEdit()
        self.stock_id.setPlaceholderText('输入商品编号（如 BH001）')
        self.stock_id.editingFinished.connect(self._on_stock_code_changed)
        self.stock_size = QComboBox()
        self.stock_size.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.stock_size.setMinimumWidth(120)
        self.stock_size.view().setMinimumWidth(200)
        self.stock_size.addItem('均码')
        # 点击下拉框时先刷新尺码列表，避免 editingFinished 时序问题导致列表为空
        _orig_showPopup = self.stock_size.showPopup
        def _refreshed_showPopup():
            prev_data = self.stock_size.currentData()
            prev_text = self.stock_size.currentText()
            self._on_stock_code_changed()
            # 还原之前的选择
            if prev_data:
                for i in range(self.stock_size.count()):
                    if self.stock_size.itemData(i) == prev_data:
                        self.stock_size.setCurrentIndex(i)
                        break
            elif prev_text:
                idx = self.stock_size.findText(prev_text)
                if idx >= 0:
                    self.stock_size.setCurrentIndex(idx)
            _orig_showPopup()
        self.stock_size.showPopup = _refreshed_showPopup
        self.stock_department = QLineEdit()
        self.stock_department.setPlaceholderText('输入部门')
        self.stock_team = QLineEdit()
        self.stock_team.setPlaceholderText('输入团名')
        self.stock_host = QLineEdit()
        self.stock_host.setPlaceholderText('主持/运营')
        self.stock_anchor = QLineEdit()
        self.stock_anchor.setPlaceholderText('主播艺名')
        self.stock_contact = QLineEdit()
        self.stock_contact.setPlaceholderText('联系方式')
        self.stock_date = QDateEdit()
        self.stock_date.setCalendarPopup(True)
        self.stock_date.setDate(QDate.currentDate())
        self.stock_date.setDisplayFormat('yyyy-MM-dd')
        self.stock_num = QLineEdit()
        self.stock_num.setPlaceholderText('输入数量')
        self.stock_num.setValidator(QIntValidator(1, 999999, self))
        self.stock_remark = QLineEdit()
        self.stock_remark.setPlaceholderText('备注信息')

        # 图片上传
        self._stock_image_path = ''
        img_btn_row = QHBoxLayout()
        img_btn_row.setSpacing(6)
        self.stock_img_btn = QPushButton('📷 选择图片')
        self.stock_img_btn.setObjectName('secondaryBtn')
        self.stock_img_btn.clicked.connect(self._select_stock_image)
        self.stock_img_clear_btn = QPushButton('✕')
        self.stock_img_clear_btn.setFixedWidth(30)
        self.stock_img_clear_btn.setObjectName('secondaryBtn')
        self.stock_img_clear_btn.clicked.connect(self._clear_stock_image)
        img_btn_row.addWidget(self.stock_img_btn)
        img_btn_row.addWidget(self.stock_img_clear_btn)
        img_btn_widget = QWidget()
        img_btn_widget.setLayout(img_btn_row)

        self.stock_img_preview = QLabel()
        self.stock_img_preview.setFixedSize(280, 180)
        self.stock_img_preview.setAlignment(Qt.AlignCenter)
        self.stock_img_preview.setStyleSheet(
            'QLabel { border: 1px dashed #c4a882; border-radius: 6px; '
            'background: #faf7f4; color: #aaa; font-size: 12px; }')
        self.stock_img_preview.setText('暂无图片')
        self.stock_img_preview.setCursor(Qt.PointingHandCursor)
        self.stock_img_preview.mousePressEvent = self._preview_stock_image

        form.addRow('编号：', self._wrap_field(self.stock_id))
        form.addRow('尺码：', self.stock_size)
        form.addRow('部门：', self.stock_department)
        form.addRow('团名：', self.stock_team)
        form.addRow('主持/运营：', self.stock_host)
        form.addRow('主播艺名：', self._wrap_field(self.stock_anchor))
        form.addRow('联系方式：', self._wrap_field(self.stock_contact))
        form.addRow('操作日期：', self.stock_date)
        form.addRow('数量：', self._wrap_field(self.stock_num))
        form.addRow('备注：', self.stock_remark)
        form.addRow('图片：', img_btn_widget)
        form_outer.addLayout(form)

        form_outer.addWidget(self.stock_img_preview)

        form_outer.addSpacing(6)

        hint = QLabel('💡 借出会自动减库存，入库会自动加库存')
        hint.setStyleSheet('color: #b08968; font-size: 11px;')
        hint.setWordWrap(True)
        form_outer.addWidget(hint)

        # 借出/归还按钮（一组）
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        self.out_btn = QPushButton('📤  借出')
        self.out_btn.setObjectName('stockOutBtn')
        self.return_btn = QPushButton('↩ 归还')
        self.return_btn.setObjectName('returnBtn')
        self.out_btn.clicked.connect(lambda: self.operate_stock(-1))
        self.return_btn.clicked.connect(self.return_stock)
        btn_row.addWidget(self.out_btn)
        btn_row.addWidget(self.return_btn)
        form_outer.addLayout(btn_row)

        form_outer.addSpacing(4)

        # 入库/赔付按钮（一组）
        btn_row2 = QHBoxLayout()
        btn_row2.setSpacing(12)
        self.in_btn = QPushButton('📥  入库')
        self.in_btn.setObjectName('stockInBtn')
        self.compensate_btn = QPushButton('💰 赔付')
        self.compensate_btn.setObjectName('compensateBtn')
        self.in_btn.clicked.connect(lambda: self.operate_stock(1))
        self.compensate_btn.clicked.connect(self.compensate_stock)
        btn_row2.addWidget(self.in_btn)
        btn_row2.addWidget(self.compensate_btn)
        form_outer.addLayout(btn_row2)

        # 修改/清空按钮（一组）
        btn_row3 = QHBoxLayout()
        btn_row3.setSpacing(12)
        self.stock_edit_btn = QPushButton('✏ 修改记录')
        self.stock_edit_btn.setObjectName('secondaryBtn')
        self.stock_edit_btn.clicked.connect(self.edit_stock_record)
        self.stock_clear_btn = QPushButton('↺ 清空表单')
        self.stock_clear_btn.setObjectName('secondaryBtn')
        self.stock_clear_btn.clicked.connect(self.clear_stock_form)
        btn_row3.addWidget(self.stock_edit_btn)
        btn_row3.addWidget(self.stock_clear_btn)
        form_outer.addLayout(btn_row3)

        # 删除记录按钮
        self.stock_del_btn = QPushButton('✕ 删除记录')
        self.stock_del_btn.setObjectName('dangerBtn')
        self.stock_del_btn.clicked.connect(self.delete_stock_record)
        form_outer.addWidget(self.stock_del_btn)

        form_outer.addStretch()

        scroll.setWidget(left_panel)

        # ── 右侧记录表 ──
        right_panel = QVBoxLayout()
        right_panel.setSpacing(10)

        # 搜索栏
        search_row = QHBoxLayout()
        self.stock_search_input = QLineEdit()
        self.stock_search_input.setObjectName('searchInput')
        self.stock_search_input.setPlaceholderText('🔍  搜索商品名 / 部门 / 团名 / 主播 ...')
        self.stock_search_input.textChanged.connect(self.search_stock_records)
        stock_export_btn = QPushButton('📊 导出Excel')
        stock_export_btn.setObjectName('secondaryBtn')
        stock_export_btn.setMinimumWidth(100)
        stock_export_btn.clicked.connect(self._export_stock_excel)
        search_row.addWidget(self.stock_search_input)
        search_row.addWidget(stock_export_btn)
        right_panel.addLayout(search_row)

        # 记录表格
        self.stock_table = QTableWidget()
        self.stock_table.setColumnCount(len(self.STOCK_RECORD_COLS))
        self.stock_table.setHorizontalHeaderLabels(self.STOCK_RECORD_COLS)
        self.stock_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.stock_table.setSelectionMode(QTableWidget.SingleSelection)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.stock_table.setAlternatingRowColors(True)
        self.stock_table.setStyleSheet('QTableWidget { alternate-background-color: #faf7f4; }')
        self.stock_table.horizontalHeader().setStretchLastSection(True)
        self.stock_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.stock_table.horizontalHeader().setMinimumSectionSize(55)
        self.stock_table.resizeColumnsToContents()
        self.stock_table.verticalHeader().setVisible(False)
        self.stock_table.setSortingEnabled(True)
        self.stock_table.horizontalHeader().setSortIndicatorShown(False)
        self.stock_table.horizontalHeader().sectionClicked.connect(
            lambda col: self._update_sort_indicator(self.stock_table, self.STOCK_RECORD_COLS, col))
        self.stock_table.cellClicked.connect(self.select_stock_record)
        right_panel.addWidget(self.stock_table, 1)

        # ── 分页栏 ──
        self._stock_page = 1
        self._stock_page_size = 50
        self._stock_data = []
        pager2 = QHBoxLayout()
        pager2.setSpacing(8)
        self.st_page_label = QLabel('共 0 条')
        self.st_page_label.setStyleSheet('color: #888; font-size: 12px;')
        pager2.addWidget(self.st_page_label)
        pager2.addStretch()
        self.st_first_btn = QPushButton('«')
        self.st_prev_btn = QPushButton('‹')
        self.st_page_info = QLabel('1 / 1')
        self.st_page_info.setStyleSheet('color: #6b5b4e; font-size: 12px; font-weight: bold;')
        self.st_page_info.setAlignment(Qt.AlignCenter)
        self.st_page_info.setMinimumWidth(60)
        self.st_next_btn = QPushButton('›')
        self.st_last_btn = QPushButton('»')
        for b in [self.st_first_btn, self.st_prev_btn, self.st_next_btn, self.st_last_btn]:
            b.setObjectName('secondaryBtn')
            b.setFixedSize(32, 28)
            b.setStyleSheet('QPushButton { padding: 0; font-size: 14px; }')
        self.st_first_btn.clicked.connect(lambda: self._stock_go_page(1))
        self.st_prev_btn.clicked.connect(lambda: self._stock_go_page(self._stock_page - 1))
        self.st_next_btn.clicked.connect(lambda: self._stock_go_page(self._stock_page + 1))
        self.st_last_btn.clicked.connect(lambda: self._stock_go_page(self._stock_total_pages()))
        pager2.addWidget(self.st_first_btn)
        pager2.addWidget(self.st_prev_btn)
        pager2.addWidget(self.st_page_info)
        pager2.addWidget(self.st_next_btn)
        pager2.addWidget(self.st_last_btn)
        pager2.addStretch()
        lbl3 = QLabel('每页')
        lbl3.setStyleSheet('color: #888; font-size: 12px;')
        self.st_page_size_combo = QComboBox()
        self.st_page_size_combo.addItems(['20', '50', '100', '200'])
        self.st_page_size_combo.setCurrentText('50')
        self.st_page_size_combo.setFixedWidth(65)
        self.st_page_size_combo.currentTextChanged.connect(self._stock_page_size_changed)
        lbl4 = QLabel('条')
        lbl4.setStyleSheet('color: #888; font-size: 12px;')
        pager2.addWidget(lbl3)
        pager2.addWidget(self.st_page_size_combo)
        pager2.addWidget(lbl4)
        right_panel.addLayout(pager2)

        layout.addWidget(scroll)
        layout.addLayout(right_panel, 1)
        return widget

    # ---------- 数据操作 ----------
    def load_clothing(self):
        data = self.db.get_all_clothing()
        self._clothing_data = data
        self._clothing_page = 1
        self._render_clothing_page()
        self._update_stats()

    def _clothing_total_pages(self):
        total = len(self._clothing_data)
        return max(1, (total + self._clothing_page_size - 1) // self._clothing_page_size)

    def _clothing_go_page(self, page):
        page = max(1, min(page, self._clothing_total_pages()))
        if page == self._clothing_page:
            return
        self._clothing_page = page
        self._render_clothing_page()

    def _clothing_page_size_changed(self, text):
        try:
            self._clothing_page_size = int(text)
        except ValueError:
            return
        self._clothing_page = 1
        self._render_clothing_page()

    def _render_clothing_page(self):
        total = len(self._clothing_data)
        pages = self._clothing_total_pages()
        self._clothing_page = max(1, min(self._clothing_page, pages))
        start = (self._clothing_page - 1) * self._clothing_page_size
        end = start + self._clothing_page_size
        page_data = self._clothing_data[start:end]
        self._fill_table(page_data)
        self.cl_page_label.setText(f'共 {total} 条')
        self.cl_page_info.setText(f'{self._clothing_page} / {pages}')
        self.cl_first_btn.setEnabled(self._clothing_page > 1)
        self.cl_prev_btn.setEnabled(self._clothing_page > 1)
        self.cl_next_btn.setEnabled(self._clothing_page < pages)
        self.cl_last_btn.setEnabled(self._clothing_page < pages)

    def _fill_table(self, data):
        self.clothing_table.setSortingEnabled(False)
        self.clothing_table.setRowCount(len(data))
        # 数值列：1=ID, 9=库存, 10=进价（偏移+1因为 col0 是复选框）
        numeric_cols = {1, 9, 10}
        low_stock_bg = QColor('#fde8e8')
        low_stock_fg = QColor('#d96b6b')
        for row, item in enumerate(data):
            # 判断该行是否低库存（库存列 index=9，即 item[9]）
            stock_val = item[9] if len(item) > 9 else None
            is_low = isinstance(stock_val, (int, float)) and stock_val <= 1
            # 复选框列
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            chk.setCheckState(Qt.Unchecked)
            if is_low:
                chk.setBackground(low_stock_bg)
            self.clothing_table.setItem(row, 0, chk)
            for col, value in enumerate(item):
                actual_col = col + 1  # 偏移 1
                # 图片列（最后一列）显示图标
                if actual_col == 11:
                    cell = SortableTableItem('📷' if value else '')
                    cell.setTextAlignment(Qt.AlignCenter)
                    self.clothing_table.setItem(row, actual_col, cell)
                    continue
                cell = SortableTableItem(str(value))
                cell.setTextAlignment(Qt.AlignCenter)
                if actual_col in numeric_cols:
                    try:
                        cell.setData(Qt.UserRole, float(value) if value else 0)
                    except (ValueError, TypeError):
                        cell.setData(Qt.UserRole, 0)
                if is_low:
                    cell.setBackground(low_stock_bg)
                    # 库存列额外红字加粗
                    if actual_col == 9:
                        cell.setForeground(low_stock_fg)
                        cell.setFont(QFont("", -1, QFont.Bold))
                self.clothing_table.setItem(row, actual_col, cell)
        self.clothing_table.resizeColumnsToContents()
        self.clothing_table.setColumnWidth(0, 36)
        self.clothing_table.setSortingEnabled(True)
        self._update_batch_count()

    def search_clothing(self, text):
        keyword = text.strip()
        data = self.db.search_clothing(keyword) if keyword else self.db.get_all_clothing()
        self._clothing_data = data
        self._clothing_page = 1
        self._render_clothing_page()

    def select_clothing(self, row, col=None):
        # 点击复选框列时不填充表单，仅更新计数
        if col == 0:
            self._update_batch_count()
            return
        table = self.clothing_table
        self.c_id.setText(table.item(row, 1).text())
        self.c_code.setText(table.item(row, 2).text())
        self.c_name.setText(table.item(row, 3).text())
        # 设置下拉框（支持自定义值）
        cat = table.item(row, 4).text()
        idx = self.c_category.findText(cat)
        if idx >= 0:
            self.c_category.setCurrentIndex(idx)
        else:
            self.c_category.setCurrentText(cat)
        self.c_brand.setText(table.item(row, 5).text())
        sz = table.item(row, 6).text()
        idx = self.c_size.findText(sz)
        if idx >= 0:
            self.c_size.setCurrentIndex(idx)
        else:
            self.c_size.setCurrentText(sz)
        self.c_color.setText(table.item(row, 7).text())
        se = table.item(row, 8).text()
        idx = self.c_season.findText(se)
        if idx >= 0:
            self.c_season.setCurrentIndex(idx)
        else:
            self.c_season.setCurrentText(se)
        self.c_stock.setText(table.item(row, 9).text())
        self.c_cost.setText(table.item(row, 10).text())
        # 图片预览
        clothing_id = int(table.item(row, 1).text())
        self.db.cursor.execute("SELECT image_path FROM clothing WHERE id=?", (clothing_id,))
        result = self.db.cursor.fetchone()
        if result and result[0]:
            self._clothing_image_path = result[0]
            self._show_clothing_preview(result[0])
        else:
            self._clear_clothing_image()

    def clear_form(self):
        for w in [self.c_id, self.c_code, self.c_name, self.c_brand, self.c_color, self.c_stock, self.c_cost]:
            w.clear()
        self.c_category.setCurrentIndex(0)
        self.c_size.setCurrentIndex(0)
        self.c_season.setCurrentIndex(0)
        self._clear_all_field_errors()
        self._clear_clothing_image()

    # ---------- 商品图片 ----------
    def _select_clothing_image(self):
        path, _ = QFileDialog.getOpenFileName(
            self, '选择商品图片', '',
            '图片文件 (*.png *.jpg *.jpeg *.bmp *.gif *.webp)')
        if path:
            self._clothing_image_path = path
            self._show_clothing_preview(path)

    def _clear_clothing_image(self):
        self._clothing_image_path = ''
        self.c_img_preview.setPixmap(QPixmap())
        self.c_img_preview.setText('暂无图片')

    def _show_clothing_preview(self, path):
        if path and os.path.isfile(path):
            pixmap = QPixmap(path)
            if not pixmap.isNull():
                scaled = pixmap.scaled(
                    self.c_img_preview.size(),
                    Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.c_img_preview.setPixmap(scaled)
                self.c_img_preview.setText('')
                return
        self.c_img_preview.setPixmap(QPixmap())
        self.c_img_preview.setText('暂无图片')

    def _preview_clothing_image(self, event=None):
        path = self._clothing_image_path
        if not path or not os.path.isfile(path):
            return
        dlg = QDialog(self)
        dlg.setWindowTitle('查看商品图片')
        dlg.setMinimumSize(600, 500)
        lay = QVBoxLayout(dlg)
        lbl = QLabel()
        lbl.setAlignment(Qt.AlignCenter)
        pixmap = QPixmap(path)
        if not pixmap.isNull():
            scaled = pixmap.scaled(580, 460, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            lbl.setPixmap(scaled)
        lay.addWidget(lbl)
        dlg.exec_()

    def _save_clothing_image(self, src_path):
        """复制图片到 clothing_images 目录，返回保存路径"""
        if not src_path or not os.path.isfile(src_path):
            return ''
        dest_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clothing_images')
        os.makedirs(dest_dir, exist_ok=True)
        ext = os.path.splitext(src_path)[1]
        filename = f"{uuid.uuid4().hex}{ext}"
        dest = os.path.join(dest_dir, filename)
        shutil.copy2(src_path, dest)
        return dest

    def _validate_clothing_form(self):
        """校验商品表单字段，返回 (ok, stock_val, cost_val)"""
        self._clear_all_field_errors()
        has_error = False
        stock_val = 0
        cost_val = 0

        name = self.c_name.text().strip()
        if not name:
            self._set_field_error(self.c_name, '请输入服装名称')
            has_error = True

        stock_text = self.c_stock.text().strip()
        if stock_text:
            try:
                stock_val = int(stock_text)
                if stock_val < 0:
                    raise ValueError
            except ValueError:
                self._set_field_error(self.c_stock, '库存数量必须为非负整数')
                has_error = True

        cost_text = self.c_cost.text().strip()
        if cost_text:
            try:
                cost_val = float(cost_text)
                if cost_val < 0:
                    raise ValueError
            except ValueError:
                self._set_field_error(self.c_cost, '进价必须为非负数字')
                has_error = True

        if has_error:
            return False, 0, 0
        return True, stock_val, cost_val

    def add_clothing(self):
        ok, stock_val, cost_val = self._validate_clothing_form()
        if not ok:
            return
        saved_img = self._save_clothing_image(self._clothing_image_path)
        self.db.add_clothing(
            self.c_code.text().strip(),
            self.c_name.text().strip(), self.c_category.currentText(),
            self.c_brand.text().strip(),
            self.c_size.currentText(), self.c_color.text().strip(),
            self.c_season.currentText(), stock_val, cost_val, saved_img)
        self.load_clothing()
        self.clear_form()
        QMessageBox.information(self, '成功', '服装商品已添加！')

    def update_clothing(self):
        if not self.c_id.text():
            QMessageBox.warning(self, '提示', '请先选择要修改的商品')
            return
        ok, stock_val, cost_val = self._validate_clothing_form()
        if not ok:
            return
        # 图片处理
        img_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clothing_images')
        if self._clothing_image_path and not self._clothing_image_path.startswith(img_dir):
            saved_img = self._save_clothing_image(self._clothing_image_path)
        else:
            saved_img = self._clothing_image_path
        self.db.update_clothing(
            int(self.c_id.text()),
            self.c_code.text().strip(),
            self.c_name.text().strip(), self.c_category.currentText(),
            self.c_brand.text().strip(), self.c_size.currentText(),
            self.c_color.text().strip(), self.c_season.currentText(),
            stock_val, cost_val, saved_img)
        self.load_clothing()
        QMessageBox.information(self, '成功', '商品信息已更新！')

    def delete_clothing(self):
        if not self.c_id.text():
            QMessageBox.warning(self, '提示', '请先选择要删除的商品')
            return
        reply = QMessageBox.question(self, '确认删除',
            f'确定要删除编号={self.c_id.text()} 吗？',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.db.delete_clothing(int(self.c_id.text()))
            self.load_clothing()
            self.clear_form()
            QMessageBox.information(self, '成功', '商品已删除！')

    # ---------- 批量操作 ----------
    def _on_clothing_header_clicked(self, col):
        if col == 0:
            self._toggle_select_all()
        else:
            self._update_sort_indicator(self.clothing_table, self.COLUMNS, col)

    def _get_checked_ids(self):
        """返回所有勾选行的 (row, clothing_id) 列表"""
        result = []
        for row in range(self.clothing_table.rowCount()):
            chk = self.clothing_table.item(row, 0)
            if chk and chk.checkState() == Qt.Checked:
                id_item = self.clothing_table.item(row, 1)
                if id_item:
                    result.append((row, int(id_item.text())))
        return result

    def _update_batch_count(self):
        count = len(self._get_checked_ids())
        self.batch_selected_label.setText(f'已选 {count} 项')
        self.batch_select_all.setText('☑ 全选' if count == self.clothing_table.rowCount() and count > 0 else '☐ 全选')

    def _toggle_select_all(self):
        checked = self._get_checked_ids()
        all_checked = len(checked) == self.clothing_table.rowCount() and self.clothing_table.rowCount() > 0
        new_state = Qt.Unchecked if all_checked else Qt.Checked
        for row in range(self.clothing_table.rowCount()):
            chk = self.clothing_table.item(row, 0)
            if chk:
                chk.setCheckState(new_state)
        self._update_batch_count()

    def _batch_delete(self):
        checked = self._get_checked_ids()
        if not checked:
            QMessageBox.warning(self, '提示', '请先勾选要删除的商品')
            return
        reply = QMessageBox.question(self, '批量删除',
            f'确定要删除选中的 {len(checked)} 件商品吗？\n\n'
            '⚠ 删除后无法恢复！',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        for _, cid in checked:
            self.db.delete_clothing(cid)
        self.load_clothing()
        self.clear_form()
        QMessageBox.information(self, '成功', f'已删除 {len(checked)} 件商品')

    def _batch_adjust_stock(self, direction):
        checked = self._get_checked_ids()
        if not checked:
            QMessageBox.warning(self, '提示', '请先勾选要调整库存的商品')
            return
        label = '增加' if direction > 0 else '减少'
        amount, ok = QInputDialog.getInt(
            self, f'批量{label}库存',
            f'已选中 {len(checked)} 件商品\n\n请输入{label}的库存数量：',
            1, 1, 999999)
        if not ok:
            return
        success = 0
        fail = 0
        for _, cid in checked:
            if self.db.update_stock(cid, amount * direction):
                success += 1
            else:
                fail += 1
        self.load_clothing()
        msg = f'已{label} {success} 件商品的库存（各 {amount} 件）'
        if fail:
            msg += f'\n{fail} 件因库存不足而跳过'
        QMessageBox.information(self, '成功', msg)

    # ---------- Excel 导入导出 ----------
    def import_excel(self):
        """从 Excel 文件批量导入商品"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, '缺少依赖', '请先安装 openpyxl：\npip install openpyxl')
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择 Excel 文件', '',
            'Excel 文件 (*.xlsx *.xls);;所有文件 (*)')
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            # 读取表头，建立列名到索引的映射
            headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col_map = {}
            header_aliases = {
                '编号': 'code', '服装编号': 'code', '商品编号': 'code', 'code': 'code',
                '名称': 'name', '服装名称': 'name', '商品名称': 'name', 'name': 'name',
                '分类': 'category', '品牌': 'brand', '尺码': 'size',
                '颜色': 'color', '季节': 'season', '库存': 'stock',
                '进价': 'cost_price', '进价(¥)': 'cost_price',
            }
            for i, h in enumerate(headers):
                key = header_aliases.get(h)
                if key:
                    col_map[key] = i

            if 'name' not in col_map:
                QMessageBox.warning(self, '格式错误',
                    '未找到"名称"或"服装名称"列。\n\n'
                    '请确保 Excel 表头包含：编号、服装名称\n'
                    '可选列：分类、品牌、尺码、颜色、季节、库存、进价')
                wb.close()
                return

            success = 0
            errors = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(v is None for v in row):
                    continue
                try:
                    def get_val(key, default=''):
                        idx = col_map.get(key)
                        if idx is not None and idx < len(row) and row[idx] is not None:
                            return row[idx]
                        return default

                    name = str(get_val('name', '')).strip()
                    if not name:
                        continue

                    code = str(get_val('code', '')).strip()
                    category = str(get_val('category', '其他')).strip()
                    brand = str(get_val('brand', '')).strip()
                    size = str(get_val('size', '均码')).strip()
                    color = str(get_val('color', '')).strip()
                    season = str(get_val('season', '四季通用')).strip()
                    stock = int(float(get_val('stock', 0)))
                    cost_price = float(get_val('cost_price', 0))

                    self.db.add_clothing(code, name, category, brand, size, color, season, stock, cost_price)
                    success += 1
                except Exception as e:
                    errors.append(f'第{row_idx}行: {e}')

            wb.close()
            self.load_clothing()

            msg = f'成功导入 {success} 条商品记录！'
            if errors:
                msg += f'\n\n{len(errors)} 条失败：\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    msg += f'\n... 还有 {len(errors)-10} 条错误'
            QMessageBox.information(self, '导入完成', msg)

        except Exception as e:
            QMessageBox.critical(self, '导入失败', f'读取 Excel 文件时出错：\n{e}')

    def export_template(self):
        """导出 Excel 导入模板"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, '缺少依赖', '请先安装 openpyxl：\npip install openpyxl')
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, '保存导入模板', '商品导入模板.xlsx',
            'Excel 文件 (*.xlsx);;所有文件 (*)')
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '商品导入'

            # 表头
            headers = ['编号', '服装名称', '分类', '品牌', '尺码', '颜色', '季节', '库存', '进价']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            # 示例数据
            sample = ['BH001', '春季新款T恤', '上衣', '优衣库', 'L', '白色', '春季', 100, 49.9]
            for col, v in enumerate(sample, 1):
                ws.cell(row=2, column=col, value=v)

            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

            wb.save(file_path)
            wb.close()
            QMessageBox.information(self, '成功', f'导入模板已保存到：\n{file_path}')
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'保存模板时出错：\n{e}')

    def _on_stock_code_changed(self):
        """当编号输入框失去焦点时，自动查询该编号对应的可借出尺码"""
        code = self.stock_id.text().strip()
        self.stock_size.clear()
        if not code:
            self.stock_size.addItem('均码')
            return
        size_rows = self.db.find_sizes_by_code(code)
        if size_rows:
            for size, stock in size_rows:
                self.stock_size.addItem(f'{size}（库存{stock}）', size)
        else:
            self.stock_size.addItem('均码')
        # 动态调整下拉列表宽度以完整显示内容
        fm = self.stock_size.fontMetrics()
        max_w = max(fm.horizontalAdvance(self.stock_size.itemText(i))
                    for i in range(self.stock_size.count())) + 40
        self.stock_size.view().setMinimumWidth(max(200, max_w))

    def operate_stock(self, direction):
        self._clear_all_field_errors()
        has_error = False

        code_text = self.stock_id.text().strip()
        if not code_text:
            self._set_field_error(self.stock_id, '请输入商品编号')
            has_error = True

        num_text = self.stock_num.text().strip()
        quantity = 0
        if not num_text:
            self._set_field_error(self.stock_num, '请输入数量')
            has_error = True
        else:
            try:
                quantity = abs(int(num_text))
                if quantity <= 0:
                    raise ValueError
            except ValueError:
                self._set_field_error(self.stock_num, '数量必须为正整数')
                has_error = True

        # 借出时校验借用人信息
        if direction == -1:
            anchor = self.stock_anchor.text().strip()
            contact = self.stock_contact.text().strip()
            if not anchor and not contact:
                self._set_field_error(self.stock_anchor, '借出请至少填写艺名或联系方式')
                self._set_field_error(self.stock_contact, '借出请至少填写艺名或联系方式')
                has_error = True
            elif contact and not all(c.isdigit() or c in '+-() ' for c in contact):
                self._set_field_error(self.stock_contact, '联系方式格式不正确')
                has_error = True

        if has_error:
            return

        # 按编号+尺码查找商品
        size_data = self.stock_size.currentData()
        size_text = size_data if size_data else self.stock_size.currentText().strip()
        result = self.db.find_clothing_by_code(code_text, size_text)
        if result is None:
            QMessageBox.warning(self, '提示', f'未找到编号为「{code_text}」尺码为「{size_text}」的商品')
            return
        clothing_id, clothing_name, current_stock = result

        num = quantity * direction
        if not self.db.update_stock(clothing_id, num):
            QMessageBox.warning(self, '失败', '库存不足，操作失败！')
            return

        op_type = '入库' if direction > 0 else '借出'
        status = '入库' if direction > 0 else '借出'
        record_date = self.stock_date.date().toString('yyyy-MM-dd')

        # 保存图片
        saved_image = self._save_stock_image(self._stock_image_path)

        self.db.add_stock_record(
            clothing_id, clothing_name,
            self.stock_department.text().strip(),
            self.stock_team.text().strip(),
            self.stock_host.text().strip(),
            self.stock_anchor.text().strip(),
            self.stock_contact.text().strip(),
            record_date, quantity, direction, status,
            self.stock_remark.text().strip(), size_text, saved_image)

        self.load_clothing()
        self.load_stock_records()
        self.clear_stock_form()
        self._update_stats()
        QMessageBox.information(self, '成功', f'{op_type} {quantity} 件完成！')

    def return_stock(self):
        """归还：选中一条借出记录，支持任意数量归还"""
        selected = self.stock_table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, '提示', '请先在右侧表格中选择一条借出记录')
            return
        record_id = int(self.stock_table.item(selected, 0).text())
        record = self.db.get_record_by_id(record_id)
        if not record:
            QMessageBox.warning(self, '提示', '未找到该记录')
            return

        current_status = record[11]
        if current_status != '借出':
            QMessageBox.warning(self, '提示', f'该记录状态为"{current_status}"，只能对"借出"状态的记录进行归还')
            return

        clothing_id = record[1]
        total_qty = record[9]

        return_qty, ok = QInputDialog.getInt(
            self, '归还数量',
            f'商品：{record[2]}（尺码：{record[14]}）\n'
            f'借出数量：{total_qty} 件\n'
            f'借用人：{record[6]}（{record[5]}）\n\n'
            f'请输入归还数量：',
            total_qty, 1, total_qty)
        if not ok:
            return

        reply = QMessageBox.question(self, '确认归还',
            f'确认归还商品：{record[2]}（尺码：{record[14]}）\n'
            f'归还数量：{return_qty} / {total_qty} 件\n'
            f'借用人：{record[6]}（{record[5]}）',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        # 库存回加
        self.db.update_stock(clothing_id, return_qty)
        self.db.split_record_for_return(record_id, return_qty)
        self.load_clothing()
        self.load_stock_records()
        self._update_stats()
        msg = f'已归还 {return_qty} 件！'
        if return_qty < total_qty:
            msg += f'\n剩余 {total_qty - return_qty} 件仍为借出状态'
        QMessageBox.information(self, '成功', msg)

    def compensate_stock(self):
        """赔付：选中借出记录，支持任意数量赔付"""
        selected = self.stock_table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, '提示', '请先在右侧表格中选择一条借出记录')
            return
        record_id = int(self.stock_table.item(selected, 0).text())
        record = self.db.get_record_by_id(record_id)
        if not record:
            QMessageBox.warning(self, '提示', '未找到该记录')
            return

        current_status = record[11]
        if current_status != '借出':
            QMessageBox.warning(self, '提示', f'该记录状态为"{current_status}"，只能对"借出"状态的记录进行赔付')
            return

        total_qty = record[9]

        # 获取商品进价，自动计算赔付金额
        cost_price = self.db.get_cost_price(record[1])

        comp_qty, ok = QInputDialog.getInt(
            self, '赔付数量',
            f'商品：{record[2]}（尺码：{record[14]}）\n'
            f'借出数量：{total_qty} 件\n'
            f'借用人：{record[6]}（{record[5]}）\n\n'
            f'请输入赔付数量：',
            total_qty, 1, total_qty)
        if not ok:
            return

        default_amount = cost_price * comp_qty
        amount, ok = QInputDialog.getDouble(
            self, '赔付金额',
            f'商品：{record[2]}（尺码：{record[14]}，赔付 {comp_qty} 件）\n'
            f'进价：¥{cost_price:.2f} × {comp_qty} 件 = ¥{default_amount:.2f}\n\n'
            f'请确认或修改赔付金额（¥）：',
            default_amount, 0, 999999, 2)
        if not ok:
            return

        reply = QMessageBox.question(self, '确认赔付',
            f'确认对以下记录进行赔付：\n'
            f'商品：{record[2]}（尺码：{record[14]}）\n'
            f'赔付数量：{comp_qty} / {total_qty} 件\n'
            f'借用人：{record[6]}（{record[5]}）\n'
            f'赔付金额：¥{amount:.2f}',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.db.split_record_for_compensate(record_id, comp_qty, amount)
        self.load_stock_records()
        self._update_stats()
        msg = f'赔付 {comp_qty} 件，¥{amount:.2f} 已记录！'
        if comp_qty < total_qty:
            msg += f'\n剩余 {total_qty - comp_qty} 件仍为借出状态'
        QMessageBox.information(self, '成功', msg)

    def delete_stock_record(self):
        """删除选中的出入库记录"""
        selected = self.stock_table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, '提示', '请先在右侧表格中选择一条记录')
            return
        record_id = int(self.stock_table.item(selected, 0).text())
        record = self.db.get_record_by_id(record_id)
        if not record:
            QMessageBox.warning(self, '提示', '未找到该记录')
            return

        status = record[11]
        clothing_id = record[1]
        quantity = record[9]
        direction = record[10]

        reply = QMessageBox.question(self, '确认删除',
            f'确认删除以下记录？\n\n'
            f'商品：{record[2]}（尺码：{record[14]}）\n'
            f'数量：{quantity} 件\n'
            f'状态：{status}\n'
            f'借用人：{record[6]}（{record[5]}）\n\n'
            f'❗ 删除后库存将自动还原',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        # 还原库存：入库记录删除后减库存，借出记录删除后加库存
        if direction == 1:
            # 入库记录 -> 删除后减库存
            self.db.update_stock(clothing_id, -quantity)
        elif status == '借出':
            # 借出未归还 -> 删除后加回库存
            self.db.update_stock(clothing_id, quantity)
        # 已归还/已赔付的记录删除不影响库存（已经还原过了）

        self.db.delete_stock_record(record_id)
        self.load_clothing()
        self.load_stock_records()
        self._update_stats()
        QMessageBox.information(self, '成功', '记录已删除')

    # ---------- 出入库记录查询 ----------
    def load_stock_records(self):
        data = self.db.get_all_stock_records()
        self._stock_data = data
        self._stock_page = 1
        self._render_stock_page()

    def _stock_total_pages(self):
        total = len(self._stock_data)
        return max(1, (total + self._stock_page_size - 1) // self._stock_page_size)

    def _stock_go_page(self, page):
        page = max(1, min(page, self._stock_total_pages()))
        if page == self._stock_page:
            return
        self._stock_page = page
        self._render_stock_page()

    def _stock_page_size_changed(self, text):
        try:
            self._stock_page_size = int(text)
        except ValueError:
            return
        self._stock_page = 1
        self._render_stock_page()

    def _render_stock_page(self):
        total = len(self._stock_data)
        pages = self._stock_total_pages()
        self._stock_page = max(1, min(self._stock_page, pages))
        start = (self._stock_page - 1) * self._stock_page_size
        end = start + self._stock_page_size
        page_data = self._stock_data[start:end]
        self._fill_stock_table(page_data)
        self.st_page_label.setText(f'共 {total} 条')
        self.st_page_info.setText(f'{self._stock_page} / {pages}')
        self.st_first_btn.setEnabled(self._stock_page > 1)
        self.st_prev_btn.setEnabled(self._stock_page > 1)
        self.st_next_btn.setEnabled(self._stock_page < pages)
        self.st_last_btn.setEnabled(self._stock_page < pages)

    def _fill_stock_table(self, data):
        self.stock_table.setSortingEnabled(False)
        self.stock_table.setRowCount(len(data))
        # 数值列索引：0=记录ID, 10=数量, 16=赔付金额
        numeric_cols = {0, 10, 16}
        for row, item in enumerate(data):
            for col, value in enumerate(item):
                # direction 列(11)显示为文字
                if col == 11:
                    display = '入库' if value == 1 else '借出'
                elif col == 16:
                    display = f'{value:.2f}' if value else '0.00'
                elif col == 17:
                    display = '📷' if value else ''
                else:
                    display = str(value) if value is not None else ''
                cell = SortableTableItem(display)
                cell.setTextAlignment(Qt.AlignCenter)
                # 数值列设置排序数据
                if col in numeric_cols:
                    try:
                        cell.setData(Qt.UserRole, float(value) if value else 0)
                    except (ValueError, TypeError):
                        cell.setData(Qt.UserRole, 0)
                # 状态列(12)着色
                if col == 12:
                    if value == '借出':
                        cell.setForeground(QColor('#d96b6b'))
                        cell.setFont(QFont("", -1, QFont.Bold))
                    elif value == '已归还':
                        cell.setForeground(QColor('#5a9e6f'))
                    elif value == '已赔付':
                        cell.setForeground(QColor('#c88a30'))
                        cell.setFont(QFont("", -1, QFont.Bold))
                self.stock_table.setItem(row, col, cell)
        self.stock_table.resizeColumnsToContents()
        self.stock_table.setSortingEnabled(True)

    def search_stock_records(self, text):
        keyword = text.strip()
        data = self.db.search_stock_records(keyword) if keyword else self.db.get_all_stock_records()
        self._stock_data = data
        self._stock_page = 1
        self._render_stock_page()

    def select_stock_record(self, row):
        """点击记录行时，将信息填充到左侧表单（用于归还/赔付/修改操作）"""
        table = self.stock_table
        # 记住选中的记录ID，用于修改操作
        record_id_item = table.item(row, 0)
        self._selected_record_id = int(record_id_item.text()) if record_id_item else None
        if table.item(row, 1):
            self.stock_id.setText(table.item(row, 1).text())
        # 尺码列 = 3
        if table.item(row, 3):
            size_text = table.item(row, 3).text()
            idx = self.stock_size.findText(size_text)
            if idx >= 0:
                self.stock_size.setCurrentIndex(idx)
            else:
                self.stock_size.clear()
                self.stock_size.addItem(size_text)
        if table.item(row, 4):
            self.stock_department.setText(table.item(row, 4).text())
        if table.item(row, 5):
            self.stock_team.setText(table.item(row, 5).text())
        if table.item(row, 6):
            self.stock_host.setText(table.item(row, 6).text())
        if table.item(row, 7):
            self.stock_anchor.setText(table.item(row, 7).text())
        if table.item(row, 8):
            self.stock_contact.setText(table.item(row, 8).text())
        if table.item(row, 10):
            self.stock_num.setText(table.item(row, 10).text())
        if table.item(row, 15):
            self.stock_remark.setText(table.item(row, 15).text())
        # 图片预览
        record_id_item = table.item(row, 0)
        if record_id_item:
            record = self.db.get_record_by_id(int(record_id_item.text()))
            if record and record[15]:
                self._show_image_preview(record[15])
            else:
                self._clear_stock_image()

    # ---------- 修改记录 ----------
    def edit_stock_record(self):
        """修改选中的出入库记录"""
        record_id = getattr(self, '_selected_record_id', None)
        if not record_id:
            QMessageBox.warning(self, '提示', '请先在右侧表格中选择一条记录')
            return
        record = self.db.get_record_by_id(record_id)
        if not record:
            QMessageBox.warning(self, '提示', '未找到该记录')
            return

        old_qty = record[9]
        old_direction = record[10]
        old_clothing_id = record[1]

        # 从表单读取新值
        self._clear_all_field_errors()
        num_text = self.stock_num.text().strip()
        if not num_text:
            self._set_field_error(self.stock_num, '请输入数量')
            return
        try:
            new_qty = abs(int(num_text))
        except ValueError:
            self._set_field_error(self.stock_num, '数量必须为正整数')
            return
        if new_qty <= 0:
            self._set_field_error(self.stock_num, '数量必须大于 0')
            return

        new_dept = self.stock_department.text().strip()
        new_team = self.stock_team.text().strip()
        new_host = self.stock_host.text().strip()
        new_anchor = self.stock_anchor.text().strip()
        new_contact = self.stock_contact.text().strip()
        new_date = self.stock_date.date().toString('yyyy-MM-dd')
        new_remark = self.stock_remark.text().strip()
        size_data = self.stock_size.currentData()
        new_size = size_data if size_data else self.stock_size.currentText().strip()

        # 图片处理：如果用户选了新图片则保存，否则保持原有
        image_path_arg = None
        if self._stock_image_path and os.path.isfile(self._stock_image_path):
            # 用户选了新文件（非已保存路径），则保存
            if not self._stock_image_path.startswith(
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'stock_images')):
                image_path_arg = self._save_stock_image(self._stock_image_path)
            else:
                image_path_arg = self._stock_image_path

        changes = []
        if new_qty != old_qty:
            changes.append(f'数量：{old_qty} → {new_qty}')
        if new_dept != record[3]:
            changes.append(f'部门：{record[3]} → {new_dept}')
        if new_team != record[4]:
            changes.append(f'团名：{record[4]} → {new_team}')
        if new_host != record[5]:
            changes.append(f'主持/运营：{record[5]} → {new_host}')
        if new_anchor != record[6]:
            changes.append(f'主播艺名：{record[6]} → {new_anchor}')
        if new_contact != record[7]:
            changes.append(f'联系方式：{record[7]} → {new_contact}')
        if new_date != record[8]:
            changes.append(f'日期：{record[8]} → {new_date}')
        if new_remark != record[13]:
            changes.append(f'备注：{record[13]} → {new_remark}')
        if new_size != record[14]:
            changes.append(f'尺码：{record[14]} → {new_size}')
        if image_path_arg is not None and image_path_arg != record[15]:
            changes.append('图片已更新')

        if not changes:
            QMessageBox.information(self, '提示', '未检测到修改')
            return

        reply = QMessageBox.question(self, '确认修改',
            f'确认修改记录 #{record_id}？\n\n'
            f'商品：{record[2]}（尺码：{record[14]}）\n\n'
            + '\n'.join(changes),
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        # 数量变化时调整库存
        if new_qty != old_qty:
            diff = (new_qty - old_qty) * old_direction
            # 借出(direction=-1)增加数量 → 库存多扣；入库(direction=1)增加数量 → 库存多加
            if not self.db.update_stock(old_clothing_id, diff):
                QMessageBox.warning(self, '失败', '库存不足，无法修改数量！')
                return

        self.db.update_record_fields(
            record_id, new_dept, new_team, new_host, new_anchor, new_contact,
            new_date, new_qty, new_remark, new_size, image_path_arg)

        self.load_clothing()
        self.load_stock_records()
        self._update_stats()
        QMessageBox.information(self, '成功', '记录已修改')

    # ---------- 图片相关 ----------
    def _select_stock_image(self):
        path, _ = QFileDialog.getOpenFileName(
            self, '选择图片', '',
            '图片文件 (*.png *.jpg *.jpeg *.bmp *.gif *.webp)')
        if path:
            self._stock_image_path = path
            self._show_image_preview(path)

    def _clear_stock_image(self):
        self._stock_image_path = ''
        self.stock_img_preview.setPixmap(QPixmap())
        self.stock_img_preview.setText('暂无图片')

    def _show_image_preview(self, path):
        if path and os.path.isfile(path):
            pixmap = QPixmap(path)
            if not pixmap.isNull():
                scaled = pixmap.scaled(
                    self.stock_img_preview.size(),
                    Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.stock_img_preview.setPixmap(scaled)
                self.stock_img_preview.setText('')
                return
        self.stock_img_preview.setPixmap(QPixmap())
        self.stock_img_preview.setText('暂无图片')

    def _preview_stock_image(self, event=None):
        """点击预览图放大查看"""
        # 优先从当前预览来源取路径
        path = self._stock_image_path
        if not path:
            # 尝试从选中行获取
            selected = self.stock_table.currentRow()
            if selected >= 0:
                record_id_item = self.stock_table.item(selected, 0)
                if record_id_item:
                    record = self.db.get_record_by_id(int(record_id_item.text()))
                    if record:
                        path = record[15]
        if not path or not os.path.isfile(path):
            return
        dlg = QDialog(self)
        dlg.setWindowTitle('查看图片')
        dlg.setMinimumSize(600, 500)
        lay = QVBoxLayout(dlg)
        lbl = QLabel()
        lbl.setAlignment(Qt.AlignCenter)
        pixmap = QPixmap(path)
        if not pixmap.isNull():
            scaled = pixmap.scaled(580, 460, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            lbl.setPixmap(scaled)
        lay.addWidget(lbl)
        dlg.exec_()

    def _save_stock_image(self, src_path):
        """复制图片到 stock_images 目录，返回保存路径"""
        if not src_path or not os.path.isfile(src_path):
            return ''
        dest_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'stock_images')
        os.makedirs(dest_dir, exist_ok=True)
        ext = os.path.splitext(src_path)[1]
        filename = f"{uuid.uuid4().hex}{ext}"
        dest = os.path.join(dest_dir, filename)
        shutil.copy2(src_path, dest)
        return dest

    def clear_stock_form(self):
        for w in [self.stock_id, self.stock_department, self.stock_team,
                  self.stock_host, self.stock_anchor, self.stock_contact,
                  self.stock_num, self.stock_remark]:
            w.clear()
        self.stock_size.clear()
        self.stock_size.addItem('均码')
        self.stock_date.setDate(QDate.currentDate())
        self._clear_stock_image()
        self._selected_record_id = None
        self._clear_all_field_errors()

    # ---------- 使用手册 ----------
    def _show_help_manual(self):
        dlg = QDialog(self)
        dlg.setWindowTitle('Nicloth 使用手册')
        dlg.resize(720, 560)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(0, 0, 0, 12)

        text = QTextBrowser()
        text.setOpenExternalLinks(False)
        text.setStyleSheet(
            'QTextBrowser { border: none; background: #fdfbf9; '
            'padding: 20px 28px; font-size: 13px; color: #333; line-height: 1.7; }')
        text.setHtml(self._HELP_HTML)
        lay.addWidget(text)

        close_btn = QPushButton('关闭')
        close_btn.setFixedWidth(100)
        close_btn.clicked.connect(dlg.accept)
        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_lay.addWidget(close_btn)
        btn_lay.addStretch()
        lay.addLayout(btn_lay)
        dlg.exec_()

    _HELP_HTML = '''
    <div style="font-family:'Microsoft YaHei','PingFang SC',sans-serif;">
    <h1 style="color:#b08968; text-align:center; margin-bottom:4px;">Nicloth 服装管理系统</h1>
    <p style="text-align:center; color:#999; font-size:12px; margin-top:0;">用户使用手册</p>
    <hr style="border:none; border-top:1px solid #e0d8d0; margin:12px 0;">

    <h2 style="color:#6b5b4e;">一、系统简介</h2>
    <p>Nicloth 服装管理系统是一款轻量化桌面端库存管理工具，专为服装行业打造，无需复杂配置即可上手使用。
    系统主打服装商品管理、出入库操作、库存实时统计、记录追溯等核心功能，兼顾日常库存管控、货品领用归还、破损赔付等业务场景。</p>

    <h2 style="color:#6b5b4e;">二、运行环境</h2>
    <ul>
    <li>支持系统：Windows、Linux、macOS</li>
    <li>额外配置：无需手动安装数据库，开箱即用</li>
    <li>初始账号：管理员 <b>admin</b>，密码 <b>123456</b></li>
    </ul>

    <h2 style="color:#6b5b4e;">三、快速入门</h2>
    <h3 style="color:#7a5c3e;">3.1 系统登录</h3>
    <ol>
    <li>双击启动程序进入登录界面，账号密码已默认填充</li>
    <li>点击<b>登录</b>按钮或在密码框按回车键快捷登录</li>
    <li>验证通过后自动跳转至系统主界面</li>
    </ol>
    <h3 style="color:#7a5c3e;">3.2 主界面概览</h3>
    <p>登录后主界面分为三大板块：</p>
    <ul>
    <li><b>顶部栏</b>：左侧显示系统名称，右侧展示当前用户名</li>
    <li><b>实时统计区</b>：商品总数、总库存量、低库存预警、待归还</li>
    <li><b>功能选项卡</b>：商品管理 和 出入库管理</li>
    </ul>

    <h2 style="color:#6b5b4e;">四、商品管理操作</h2>
    <h3 style="color:#7a5c3e;">4.1 信息填写规则</h3>
    <table style="border-collapse:collapse; width:100%; font-size:12px; margin:8px 0;">
    <tr style="background:#f9f5f1;">
        <th style="border:1px solid #e0d8d0; padding:6px 8px; color:#6b5b4e;">字段</th>
        <th style="border:1px solid #e0d8d0; padding:6px 8px; color:#6b5b4e;">说明</th>
    </tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">ID</td><td style="border:1px solid #e0d8d0; padding:6px 8px;">系统自动生成，只读</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">编号</td><td style="border:1px solid #e0d8d0; padding:6px 8px;">自定义货品编码，支持模糊搜索</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">名称 *</td><td style="border:1px solid #e0d8d0; padding:6px 8px;">必填，服装货品全称</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">分类</td><td style="border:1px solid #e0d8d0; padding:6px 8px;">下拉选择：上衣、裤装、裙装、外套等</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">尺码</td><td style="border:1px solid #e0d8d0; padding:6px 8px;">下拉选择：XS ~ XXXL、均码</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">库存 *</td><td style="border:1px solid #e0d8d0; padding:6px 8px;">非负整数，默认 0</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">进价</td><td style="border:1px solid #e0d8d0; padding:6px 8px;">非负数字，默认 0.00</td></tr>
    </table>
    <h3 style="color:#7a5c3e;">4.2 新增 / 修改 / 删除</h3>
    <ul>
    <li><b>新增</b>：填写表单 → 点击 ✚ 新增</li>
    <li><b>修改</b>：选中表格行 → 修改表单 → 点击 ✎ 修改</li>
    <li><b>删除</b>：选中表格行 → 点击 ✕ 删除 → 确认</li>
    </ul>
    <p style="color:#d96b6b;"><span style="color:#d96b6b;">&#9888;</span> 商品删除后无法恢复，删除前请确认无关联的未结出入库记录。</p>
    <h3 style="color:#7a5c3e;">4.3 搜索与排序</h3>
    <ul>
    <li>输入关键词即可模糊搜索名称、编号、分类、品牌、颜色</li>
    <li>点击列表头可按该列升序/降序排列</li>
    </ul>

    <h2 style="color:#6b5b4e;">五、出入库管理操作</h2>
    <h3 style="color:#7a5c3e;">5.1 货品借出（出库）</h3>
    <ol>
    <li>切换至「出入库管理」选项卡</li>
    <li>输入商品编号，选择尺码</li>
    <li>填写部门、团队、经办人、联系方式、借出数量等</li>
    <li>点击 <b>📤 借出</b> 按钮，系统自动扣减库存</li>
    </ol>
    <h3 style="color:#7a5c3e;">5.2 货品入库</h3>
    <ol>
    <li>输入商品编号、选择尺码、填写入库数量</li>
    <li>点击 <b>📥 入库</b> 按钮，系统自动增加库存</li>
    </ol>
    <h3 style="color:#7a5c3e;">5.3 货品归还</h3>
    <ol>
    <li>选中状态为「借出」的记录</li>
    <li>点击 <b>↩ 归还</b> 按钮，输入归还数量</li>
    <li>全额归还→状态改为已归还；部分归还→自动拆分记录</li>
    </ol>
    <h3 style="color:#7a5c3e;">5.4 货品赔付</h3>
    <ol>
    <li>选中借出记录，点击 <b>💰 赔付</b></li>
    <li>输入赔付数量和金额（默认按进价×数量自动计算）</li>
    <li>全额/部分赔付同归还逻辑</li>
    </ol>
    <p style="color:#d96b6b;"><span style="color:#d96b6b;">&#9888;</span> 赔付操作不会恢复库存，仅做状态标记和金额记录。</p>
    <h3 style="color:#7a5c3e;">5.5 记录修改与删除</h3>
    <ul>
    <li><b>修改</b>：选中记录 → 修改表单内容 → 点击 ✏ 修改记录（数量变化会同步调整库存）</li>
    <li><b>删除</b>：选中记录 → 点击 ✕ 删除记录 → 确认（库存自动还原）</li>
    </ul>
    <h3 style="color:#7a5c3e;">5.6 图片上传</h3>
    <p>出入库时可上传图片留档，点击「📷 选择图片」选取，点击缩略图可放大查看。</p>

    <h2 style="color:#6b5b4e;">六、数据安全与备份</h2>
    <ul>
    <li>系统每 30 分钟自动备份数据库到 <code style="background:#f5f0eb; padding:2px 6px; border-radius:4px;">backups/</code> 目录，保留最近 10 份</li>
    <li>数据库文件为 <code style="background:#f5f0eb; padding:2px 6px; border-radius:4px;">clothing_db.db</code>，存放在程序目录内</li>
    <li>建议批量操作前手动备份数据库文件</li>
    <li>请勿随意修改或删除数据库文件</li>
    </ul>

    <h2 style="color:#6b5b4e;">七、常见问题</h2>
    <table style="border-collapse:collapse; width:100%; font-size:12px; margin:8px 0;">
    <tr style="background:#f9f5f1;">
        <th style="border:1px solid #e0d8d0; padding:6px 8px; color:#6b5b4e;">问题</th>
        <th style="border:1px solid #e0d8d0; padding:6px 8px; color:#6b5b4e;">解决方法</th>
    </tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">提示库存不足</td>
        <td style="border:1px solid #e0d8d0; padding:6px 8px;">核对当前库存，调减出库数量</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">归还/赔付数量错误</td>
        <td style="border:1px solid #e0d8d0; padding:6px 8px;">输入大于0且不超出借出总数的整数</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">搜索无结果</td>
        <td style="border:1px solid #e0d8d0; padding:6px 8px;">简化关键词，只输入核心名称或编号</td></tr>
    <tr><td style="border:1px solid #e0d8d0; padding:6px 8px;">登录失败</td>
        <td style="border:1px solid #e0d8d0; padding:6px 8px;">使用 admin / 123456 登录，注意大小写</td></tr>
    </table>

    <h2 style="color:#6b5b4e;">八、注意事项</h2>
    <div style="background:#fff5f5; border-left:4px solid #d96b6b; padding:10px 14px; border-radius:4px; margin:8px 0;">
    <p style="margin:4px 0;"><span style="color:#d96b6b;">&#9888;</span> 删除商品/记录前务必再三确认，删除后<b>不可恢复</b></p>
    <p style="margin:4px 0;"><span style="color:#d96b6b;">&#9888;</span> 出入库操作直接影响库存数据，务必如实填写</p>
    <p style="margin:4px 0;"><span style="color:#d96b6b;">&#9888;</span> 仅授权管理员登录系统，禁止无关人员操作</p>
    <p style="margin:4px 0;"><span style="color:#d96b6b;">&#9888;</span> 勿随意删除程序目录内的数据库文件</p>
    </div>
    </div>
    '''

    # ---------- Excel 导出 ----------
    def _export_clothing_excel(self):
        """导出商品数据到 Excel（勾选行 / 全部）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            QMessageBox.warning(self, '缺少依赖', '请先安装 openpyxl：\npip install openpyxl')
            return

        checked = self._get_checked_ids()
        if checked:
            rows_to_export = [r for r, _ in checked]
            default_name = f'商品数据_选中{len(checked)}条.xlsx'
        else:
            rows_to_export = list(range(self.clothing_table.rowCount()))
            default_name = '商品数据_全部.xlsx'

        if not rows_to_export:
            QMessageBox.warning(self, '提示', '没有可导出的数据')
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, '导出商品数据', default_name, 'Excel 文件 (*.xlsx);;所有文件 (*)')
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '商品数据'
            headers = self.COLUMNS[1:]  # 跳过复选框列
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            for ri, row in enumerate(rows_to_export, 2):
                for ci, col in enumerate(range(1, self.clothing_table.columnCount()), 1):
                    item = self.clothing_table.item(row, col)
                    ws.cell(row=ri, column=ci, value=item.text() if item else '')
            for c in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 14
            wb.save(file_path)
            wb.close()
            QMessageBox.information(self, '成功', f'已导出 {len(rows_to_export)} 条商品数据\n{file_path}')
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'保存时出错：\n{e}')

    def _export_stock_excel(self):
        """导出出入库记录到 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            QMessageBox.warning(self, '缺少依赖', '请先安装 openpyxl：\npip install openpyxl')
            return

        row_count = self.stock_table.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, '提示', '没有可导出的数据')
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, '导出出入库记录', '出入库记录.xlsx', 'Excel 文件 (*.xlsx);;所有文件 (*)')
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '出入库记录'
            headers = list(self.STOCK_RECORD_COLS)
            headers[-1] = '图片路径'  # 图片列导出路径文字
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            for ri in range(row_count):
                for ci in range(len(headers)):
                    item = self.stock_table.item(ri, ci)
                    ws.cell(row=ri + 2, column=ci + 1, value=item.text() if item else '')
            for c in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 14
            wb.save(file_path)
            wb.close()
            QMessageBox.information(self, '成功', f'已导出 {row_count} 条记录\n{file_path}')
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'保存时出错：\n{e}')

    # ---------- 系统设置标签 ----------
    def _create_settings_tab(self):
        widget = QWidget()
        outer = QVBoxLayout(widget)
        outer.setContentsMargins(40, 30, 40, 30)
        outer.setSpacing(24)

        title = QLabel('🛠  系统设置')
        title.setStyleSheet('font-size: 18px; font-weight: bold; color: #6b5b4e;')
        outer.addWidget(title)

        # ── 数据备份与恢复 ──
        group = QWidget()
        group.setObjectName('settingsCard')
        group.setStyleSheet(
            '#settingsCard { background: #fff; border: 1px solid #e0d5ca; '
            'border-radius: 10px; padding: 20px; color: #555; }')
        g_layout = QVBoxLayout(group)
        g_layout.setSpacing(14)

        g_title = QLabel('💾  数据备份与恢复')
        g_title.setStyleSheet('font-size: 14px; font-weight: bold; color: #6b5b4e;')
        g_layout.addWidget(g_title)

        desc = QLabel(
            '手动备份会将当前数据库复制到您选择的位置。\n'
            '恢复操作会用选定的备份文件替换当前数据库，请谨慎操作。')
        desc.setStyleSheet('color: #888; font-size: 12px;')
        desc.setWordWrap(True)
        g_layout.addWidget(desc)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(14)
        backup_btn = QPushButton('📦  手动备份')
        backup_btn.setObjectName('secondaryBtn')
        backup_btn.setFixedHeight(36)
        backup_btn.clicked.connect(self._manual_backup)
        restore_btn = QPushButton('📂  恢复数据')
        restore_btn.setObjectName('secondaryBtn')
        restore_btn.setFixedHeight(36)
        restore_btn.clicked.connect(self._restore_backup)
        btn_row.addWidget(backup_btn)
        btn_row.addWidget(restore_btn)
        btn_row.addStretch()
        g_layout.addLayout(btn_row)

        self._backup_status = QLabel('')
        self._backup_status.setStyleSheet('color: #888; font-size: 11px;')
        g_layout.addWidget(self._backup_status)

        outer.addWidget(group)

        # ── 自动备份信息 ──
        auto_group = QWidget()
        auto_group.setObjectName('settingsCard2')
        auto_group.setStyleSheet(
            '#settingsCard2 { background: #fff; border: 1px solid #e0d5ca; '
            'border-radius: 10px; padding: 20px; color: #555; }')
        ag_layout = QVBoxLayout(auto_group)
        ag_layout.setSpacing(10)
        ag_title = QLabel('🔄  自动备份')
        ag_title.setStyleSheet('font-size: 14px; font-weight: bold; color: #6b5b4e;')
        ag_layout.addWidget(ag_title)
        auto_desc = QLabel(
            '系统每 30 分钟自动备份一次数据库到 backups/ 目录，保留最近 10 个备份。\n'
            '启动时也会自动执行一次备份。')
        auto_desc.setStyleSheet('color: #888; font-size: 12px;')
        auto_desc.setWordWrap(True)
        ag_layout.addWidget(auto_desc)
        outer.addWidget(auto_group)

        outer.addStretch()
        return widget

    def _manual_backup(self):
        """手动备份：复制 db 文件到用户选择的位置"""
        db_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'clothing_db.db')
        if not os.path.isfile(db_path):
            QMessageBox.warning(self, '提示', '未找到数据库文件')
            return
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f'clothing_db_backup_{ts}.db'
        file_path, _ = QFileDialog.getSaveFileName(
            self, '选择备份保存位置', default_name,
            '数据库文件 (*.db);;所有文件 (*)')
        if not file_path:
            return
        try:
            shutil.copy2(db_path, file_path)
            self._backup_status.setText(f'✅ 备份成功：{file_path}')
            self._backup_status.setStyleSheet('color: #2e7d32; font-size: 11px;')
        except Exception as e:
            QMessageBox.critical(self, '备份失败', f'备份出错：\n{e}')

    def _restore_backup(self):
        """恢复数据：用选定的 db 文件替换当前数据库"""
        reply = QMessageBox.warning(
            self, '确认恢复',
            '恢复操作会覆盖当前所有数据，建议先手动备份！\n\n确定要继续吗？',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择备份文件', '',
            '数据库文件 (*.db);;所有文件 (*)')
        if not file_path:
            return
        db_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'clothing_db.db')
        try:
            # 先关闭当前连接
            self.db.conn.close()
            shutil.copy2(file_path, db_path)
            # 重新打开连接
            self.db.conn = sqlite3.connect(db_path)
            self.db.cursor = self.db.conn.cursor()
            self.load_clothing()
            self.load_stock_records()
            self._update_stats()
            self._backup_status.setText(f'✅ 数据已恢复：{os.path.basename(file_path)}')
            self._backup_status.setStyleSheet('color: #2e7d32; font-size: 11px;')
            QMessageBox.information(self, '成功', '数据库已恢复，数据已刷新！')
        except Exception as e:
            # 尝试重新连接
            try:
                self.db.conn = sqlite3.connect(db_path)
                self.db.cursor = self.db.conn.cursor()
            except Exception:
                pass
            QMessageBox.critical(self, '恢复失败', f'恢复出错：\n{e}')

    # ---------- 数据自动备份 ----------
    def _setup_auto_backup(self):
        """启动时立即备份一次，然后每30分钟自动备份"""
        self._backup_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'backups')
        os.makedirs(self._backup_dir, exist_ok=True)
        self._do_backup()  # 启动时备份
        self._backup_timer = QTimer(self)
        self._backup_timer.timeout.connect(self._do_backup)
        self._backup_timer.start(30 * 60 * 1000)  # 30分钟

    def _do_backup(self):
        """执行数据库备份，保留最近10个备份文件"""
        try:
            db_path = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), 'clothing_db.db')
            if not os.path.isfile(db_path):
                return
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f'clothing_db_{ts}.db'
            dest = os.path.join(self._backup_dir, backup_name)
            shutil.copy2(db_path, dest)
            # 清理旧备份，只保留最近10个
            backups = sorted(
                [f for f in os.listdir(self._backup_dir)
                 if f.startswith('clothing_db_') and f.endswith('.db')],
                reverse=True)
            for old in backups[10:]:
                os.remove(os.path.join(self._backup_dir, old))
        except Exception:
            pass  # 备份失败不影响正常使用


# ===================== 程序入口 =====================
if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet(STYLE_SHEET)
    login = LoginWindow()
    login.show()
    sys.exit(app.exec_())